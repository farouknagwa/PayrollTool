#!/usr/bin/env python3
"""Replace the attendance period in the Final Nagwa Technologies report.

The source workbook lives in ``templates/`` and the result is written to
``output/`` using the same file name.

The period is auto-detected from the raw attendance report
(``raw data/Attendance Report.xls``): the first date in the "Attendance
Day" column decides which 21st->20th period to build. For example, a first
date of 28-04-2026 yields the period 21 April -> 20 May.

The script rewrites the date columns (``I3:AM3``) so they cover the detected
period from the 21st of one month to the 20th of the next month, and clears
any existing attendance values in the employee rows for those columns. All
other formatting, formulas and summary columns are preserved.
"""

from __future__ import annotations

import argparse
import calendar
import os
import re
import tempfile
from zipfile import ZipFile, ZIP_DEFLATED
from datetime import date, datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Paths are resolved relative to the project root (the parent of scripts/).
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
DEFAULT_INPUT = os.path.join(
    PROJECT_ROOT, "templates", "Final Nagwa Technologies.xlsx"
)
DEFAULT_OUTPUT_DIR = os.path.join(PROJECT_ROOT, "output")
DEFAULT_REPORT = os.path.join(PROJECT_ROOT, "raw data", "Attendance Report.xls")

DATE_ROW = 3
HEADER_ROW = 4
FIRST_DATA_ROW = 5
DATE_COL_START = 9   # column I
DATE_COL_END = 39    # column AM (inclusive) -> 31 columns total


def effective_last_employee_row(sheet, first_row: int = FIRST_DATA_ROW, id_col: int = 1) -> int:
    """Return the last employee row, ignoring formatting-only trailing rows."""
    last = first_row - 1
    blank_run = 0
    for row in range(first_row, sheet.max_row + 1):
        if sheet.cell(row, id_col).value is None:
            if last >= first_row:
                blank_run += 1
                if blank_run >= 50:
                    break
            continue
        last = row
        blank_run = 0
    return max(last, first_row - 1)


def trim_formatting_only_rows(sheet, last_row: int) -> None:
    """Drop worksheet cell records beyond the employee block without shifting rows."""
    for coord in [coord for coord in sheet._cells if coord[0] > last_row]:
        del sheet._cells[coord]
    for row_index in [row for row in sheet.row_dimensions if row > last_row]:
        del sheet.row_dimensions[row_index]


ROW_RE = re.compile(rb"<row\b[^>]*\br=\"(\d+)\"[^>]*>.*?</row>")


def row_has_column_value(row_xml: bytes, row_num: int, col_letter: str = "A") -> bool:
    cell_re = re.compile(
        rb"<c\b[^>]*\br=\"" + col_letter.encode("ascii") + str(row_num).encode("ascii") + rb"\"[^>]*>(.*?)</c>"
    )
    match = cell_re.search(row_xml)
    if not match:
        return False
    body = match.group(1)
    return b"<v>" in body or b"<is>" in body


def compact_template_rows(path: str, first_row: int = FIRST_DATA_ROW) -> str:
    """Create a compact copy of the workbook without formatting-only tail rows."""
    with ZipFile(path, "r") as source:
        sheet_name = "xl/worksheets/sheet1.xml"
        if sheet_name not in source.namelist():
            return path
        sheet_xml = source.read(sheet_name)

        last_employee_row = first_row - 1
        blank_run = 0
        for match in ROW_RE.finditer(sheet_xml):
            row_num = int(match.group(1))
            if row_num < first_row:
                continue
            if row_has_column_value(match.group(0), row_num):
                last_employee_row = row_num
                blank_run = 0
            elif last_employee_row >= first_row:
                blank_run += 1
                if blank_run >= 50:
                    break

        if last_employee_row < first_row:
            return path

        rows = []
        max_seen_row = 0
        for match in ROW_RE.finditer(sheet_xml):
            row_num = int(match.group(1))
            max_seen_row = max(max_seen_row, row_num)
            if row_num < first_row or row_num <= last_employee_row:
                rows.append(match.group(0))

        if max_seen_row <= last_employee_row:
            return path

        sheet_data_match = re.search(rb"(<sheetData>).*?(</sheetData>)", sheet_xml)
        if not sheet_data_match:
            return path

        compact_xml = (
            sheet_xml[: sheet_data_match.start()]
            + sheet_data_match.group(1)
            + b"".join(rows)
            + sheet_data_match.group(2)
            + sheet_xml[sheet_data_match.end() :]
        )

        def replace_dimension(match):
            ref = match.group(2).decode("ascii")
            end_ref = ref.split(":")[-1]
            end_col_match = re.match(r"([A-Z]+)", end_ref)
            end_col = end_col_match.group(1) if end_col_match else "A"
            return (
                match.group(1)
                + f"A1:{end_col}{last_employee_row}".encode("ascii")
                + match.group(3)
            )

        compact_xml = re.sub(
            rb"(<dimension ref=\")([^\"]+)(\"/?>)",
            replace_dimension,
            compact_xml,
            count=1,
        )

        fd, compact_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        with ZipFile(compact_path, "w", ZIP_DEFLATED) as target:
            for item in source.infolist():
                data = compact_xml if item.filename == sheet_name else source.read(item.filename)
                target.writestr(item, data)
        print(
            f"  Compacted template rows from {max_seen_row} to {last_employee_row}.",
            flush=True,
        )
        return compact_path


# ----------------------------------------------------------------------
# Period detection
# ----------------------------------------------------------------------

def compute_period(ref: date) -> tuple[date, date]:
    """Return the (start, end) of the HR period containing ``ref``.

    Periods run from the 21st of one month through the 20th of the next
    month, inclusive. The period chosen is the one that contains ``ref``:
      - day >= 21 -> period starts on the 21st of ``ref``'s month;
      - day <= 20 -> period starts on the 21st of the previous month.
    """
    if ref.day >= 21:
        start = date(ref.year, ref.month, 21)
    elif ref.month == 1:
        start = date(ref.year - 1, 12, 21)
    else:
        start = date(ref.year, ref.month - 1, 21)

    if start.month == 12:
        end = date(start.year + 1, 1, 20)
    else:
        end = date(start.year, start.month + 1, 20)
    return start, end


def detect_first_attendance_date(report_path: str) -> date:
    """Read the first date in the "Attendance Day" column of the report."""
    df = pd.read_excel(report_path, header=0)
    col = "Attendance Day"
    if col not in df.columns:
        raise SystemExit(
            f"Column {col!r} not found in {report_path!r}. "
            f"Available columns: {list(df.columns)}"
        )
    series = df[col].dropna()
    if series.empty:
        raise SystemExit(f"No dates found in column {col!r} of {report_path!r}.")

    first = series.iloc[0]
    if isinstance(first, str):
        return datetime.strptime(first.strip(), "%d/%m/%Y").date()
    return pd.to_datetime(first).date()


def build_period(start: date, end: date) -> list[datetime]:
    """Return the list of dates from ``start`` to ``end`` (inclusive)."""
    dates = []
    cur = start
    while cur <= end:
        dates.append(datetime(cur.year, cur.month, cur.day))
        # next day
        d = cur.day + 1
        m = cur.month
        y = cur.year
        if d > calendar.monthrange(y, m)[1]:
            d = 1
            m += 1
            if m > 12:
                m = 1
                y += 1
        cur = date(y, m, d)
    return dates


def replace_period(input_path: str, output_path: str, ref_date: date) -> None:
    start_date, end_date = compute_period(ref_date)
    dates = build_period(start_date, end_date)
    n_dates = len(dates)
    n_slots = DATE_COL_END - DATE_COL_START + 1
    if n_dates > n_slots:
        raise RuntimeError(
            f"Period has {n_dates} days but the sheet only has "
            f"{n_slots} date columns ({get_column_letter(DATE_COL_START)}:"
            f"{get_column_letter(DATE_COL_END)})."
        )

    compact_path = compact_template_rows(input_path)
    wb = load_workbook(compact_path)
    ws = wb.active
    last_row = effective_last_employee_row(ws)
    if ws.max_row > last_row:
        trim_formatting_only_rows(ws, last_row)

    # 1) Rewrite the date row (row 3). Leave trailing slots empty if the
    #    period is shorter than the available 31 columns.
    for offset in range(n_slots):
        col = DATE_COL_START + offset
        cell = ws.cell(row=DATE_ROW, column=col)
        if offset < n_dates:
            cell.value = dates[offset]
            # Preserve a date format if the cell did not already have one.
            if not cell.number_format or cell.number_format == "General":
                cell.number_format = "dd-mmm"
        else:
            cell.value = None

    # 2) Clear existing attendance values for every employee row inside the
    #    date column range. Skip the header rows (1-4) and the summary block
    #    after column AM.
    print(f"  Clearing final report values through employee row {last_row}.", flush=True)
    for row in range(FIRST_DATA_ROW, last_row + 1):
        for col in range(DATE_COL_START, DATE_COL_END + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.value = None

    wb.save(output_path)
    wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Replace the attendance period in the Final Nagwa report."
    )
    parser.add_argument(
        "-i", "--input",
        default=DEFAULT_INPUT,
        help="Source workbook (default: templates/Final Nagwa Technologies.xlsx).",
    )
    parser.add_argument(
        "-o", "--output",
        default=None,
        help=(
            "Output workbook. Defaults to the output/ folder using the "
            "same file name as the input."
        ),
    )
    parser.add_argument(
        "-r", "--report",
        default=DEFAULT_REPORT,
        help=(
            "Attendance report used to detect the period "
            "(default: raw data/Attendance Report.xls)."
        ),
    )
    args = parser.parse_args()

    inp = os.path.abspath(args.input)
    if not os.path.isfile(inp):
        parser.error(f"Input file not found: {inp}")

    report = os.path.abspath(args.report)
    if not os.path.isfile(report):
        parser.error(f"Attendance report not found: {report}")

    if args.output:
        out = os.path.abspath(args.output)
    else:
        out = os.path.join(DEFAULT_OUTPUT_DIR, os.path.basename(inp))

    os.makedirs(os.path.dirname(out), exist_ok=True)

    first_date = detect_first_attendance_date(report)
    print(f"Detected first attendance date: {first_date.isoformat()}")

    replace_period(inp, out, first_date)

    start_date, end_date = compute_period(first_date)
    dates = build_period(start_date, end_date)
    print(
        f"Wrote {len(dates)} dates "
        f"({dates[0].strftime('%d %b %Y')} - {dates[-1].strftime('%d %b %Y')}) "
        f"to '{out}'."
    )


if __name__ == "__main__":
    main()
