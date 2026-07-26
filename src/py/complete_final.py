#!/usr/bin/env python3
"""Copy per-day attendance values from the source Nagwa workbook into the
final monthly report.

Source workbook: ``Nagwa Technologies.xlsx``
    - Sheet ``Nagwa Technologies``
    - Column 1 (rows 4+) holds the employee ID.
    - Row 2 holds the date headers starting at column 10.
        * Some dates span 4 sub-columns (``in`` / ``out`` / ``Leave`` /
          ``Shortage`` in row 3).  We take the value of the ``Shortage``
          column for that day.
        * Other dates occupy a single column (no sub-headers).  We take the
          value sitting directly under the date.

Target workbook: ``Final Nagwa Technologies.xlsx``
    - Sheet ``Final Nagwa Technologies``
    - Column 1 (rows 5+) holds the employee ID.
    - Row 3 (cols I..AM) holds one date per column.

For every (employee, date) pair that exists in both workbooks the script
writes the chosen source value into the corresponding cell of the target
workbook.  Column AN (``Total``) keeps the template's array formula; it is
never overwritten with a static value. Column widths are auto-fitted to
contents before save. All other cells, formulas and formatting are left
untouched.

Usage:
    python complete_final.py [-s SOURCE.xlsx] [-t TARGET.xlsx]

Defaults:
    SOURCE = output/Nagwa Technologies.xlsx
    TARGET = output/Final Nagwa Technologies.xlsx
"""

from __future__ import annotations

import argparse
import os
import re
from datetime import datetime, date
from typing import Dict, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula


_OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "output")
SOURCE_PATH_DEFAULT = os.path.join(_OUTPUT_DIR, "Nagwa Technologies.xlsx")
TARGET_PATH_DEFAULT = os.path.join(_OUTPUT_DIR, "Final Nagwa Technologies.xlsx")

SOURCE_SHEET = "Nagwa Technologies"
TARGET_SHEET = "Final Nagwa Technologies"

SOURCE_DATE_ROW = 2
SOURCE_SUBHEADER_ROW = 3
SOURCE_FIRST_DATA_ROW = 4
SOURCE_FIRST_DATE_COL = 10

TARGET_DATE_ROW = 3
TARGET_FIRST_DATA_ROW = 5
TARGET_FIRST_DATE_COL = 9   # column I
TARGET_LAST_DATE_COL = 39   # column AM
TARGET_TOTAL_COL = 40       # column AN ('Total')


def effective_last_employee_row(sheet, first_row: int, id_col: int = 1) -> int:
    """Return the last employee row, ignoring formatting-only trailing rows."""
    last = first_row - 1
    blank_run = 0
    for row in range(first_row, sheet.max_row + 1):
        if sheet.cell(row, id_col).value is None:
            if last >= first_row:
                blank_run += 1
                if blank_run >= 50:
                    break
            continue
        last = row
        blank_run = 0
    return max(last, first_row - 1)


ABBREVIATIONS: Dict[str, str] = {
    "absent": "A",
    "unpaid leave": "UL",
    "permitted absence during probation": "PD",
    "permitted absence": "PA",
    "5 days sick leave": "5S",
    "severe illness sick leave": "SS",
    "severe sick 85%": "S",
    "unpaid sick leave": "US",
    "work from home": "WFH",
}


_DURATION_RE = re.compile(r"^\s*(\d{1,3}):([0-5]?\d)\s*$")


def format_duration(value):
    """Convert ``H:MM`` duration strings into ``H:MM:SS``.

    Leaves anything else untouched (entry/exit times like ``"9:26 AM"`` are
    not affected because of the trailing ``AM``/``PM``).
    """
    if not isinstance(value, str):
        return value
    m = _DURATION_RE.match(value)
    if not m:
        return value
    hours = int(m.group(1))
    minutes = int(m.group(2))
    return f"{hours}:{minutes:02d}:00"


def _normalise_label(text: str) -> str:
    """Lower-case and strip punctuation/whitespace for tolerant matching."""
    s = text.lower()
    s = s.replace("(", " ").replace(")", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


_ABBREV_LOOKUP = {_normalise_label(k): v for k, v in ABBREVIATIONS.items()}


def abbreviate(value):
    """Return the abbreviated form of ``value`` if it matches a known label.

    Matching is case-insensitive, tolerant of surrounding parentheses, and
    collapses internal whitespace.  Non-string values pass through unchanged.
    """
    if not isinstance(value, str):
        return value
    return _ABBREV_LOOKUP.get(_normalise_label(value), value)


def extract_date(cell_value) -> Optional[date]:
    """Pull a calendar date out of a header cell.

    Accepts ``datetime``/``date`` objects and strings such as
    ``"21/02/2026"`` or ``"21/02/2026 PUBLIC Holiday"``.
    """
    if isinstance(cell_value, datetime):
        return cell_value.date()
    if isinstance(cell_value, date):
        return cell_value
    if isinstance(cell_value, str):
        m = re.match(r"\s*(\d{1,2})/(\d{1,2})/(\d{4})", cell_value)
        if m:
            d, mo, y = (int(x) for x in m.groups())
            try:
                return date(y, mo, d)
            except ValueError:
                return None
    return None


def build_source_date_maps(sheet) -> tuple[Dict[date, int], Dict[date, int]]:
    """Return ``(merged_map, single_map)`` for the source sheet.

    ``merged_map[d]``  -> column index of the ``in`` sub-column (the
    ``Shortage`` column is that index + 3).
    ``single_map[d]``  -> column index where the date sits as a single,
    non-merged header.
    """
    merged_map: Dict[date, int] = {}
    merged_anchor: Dict[int, int] = {}

    for mc in sheet.merged_cells.ranges:
        if mc.min_row > SOURCE_DATE_ROW or mc.max_row < SOURCE_DATE_ROW:
            continue
        if mc.min_col < SOURCE_FIRST_DATE_COL:
            continue

        for c in range(mc.min_col, mc.max_col + 1):
            merged_anchor[c] = mc.min_col

        if mc.max_col - mc.min_col + 1 != 4:
            continue
        sub = sheet.cell(SOURCE_SUBHEADER_ROW, mc.min_col).value
        if not (isinstance(sub, str) and sub.strip().lower() == "in"):
            continue
        d = extract_date(sheet.cell(SOURCE_DATE_ROW, mc.min_col).value)
        if d is not None:
            merged_map[d] = mc.min_col

    single_map: Dict[date, int] = {}
    for col in range(SOURCE_FIRST_DATE_COL, sheet.max_column + 1):
        anchor = merged_anchor.get(col)
        if anchor is not None and anchor != col:
            continue
        d = extract_date(sheet.cell(SOURCE_DATE_ROW, col).value)
        if d is None or d in merged_map:
            continue
        single_map[d] = col

    return merged_map, single_map


def build_source_id_row_map(sheet) -> Dict[int, int]:
    """Map employee ID -> row number in the source sheet."""
    out: Dict[int, int] = {}
    last_row = effective_last_employee_row(sheet, SOURCE_FIRST_DATA_ROW)
    for r in range(SOURCE_FIRST_DATA_ROW, last_row + 1):
        v = sheet.cell(r, 1).value
        try:
            emp_id = int(v)
        except (TypeError, ValueError):
            continue
        out[emp_id] = r
    return out


def build_target_date_col_map(sheet) -> Dict[date, int]:
    """Map calendar date -> column index in the target sheet (row 3)."""
    out: Dict[date, int] = {}
    for col in range(TARGET_FIRST_DATE_COL, TARGET_LAST_DATE_COL + 1):
        d = extract_date(sheet.cell(TARGET_DATE_ROW, col).value)
        if d is not None:
            out[d] = col
    return out


def build_target_id_row_map(sheet) -> Dict[int, int]:
    """Map employee ID -> row number in the target sheet."""
    out: Dict[int, int] = {}
    last_row = effective_last_employee_row(sheet, TARGET_FIRST_DATA_ROW)
    for r in range(TARGET_FIRST_DATA_ROW, last_row + 1):
        v = sheet.cell(r, 1).value
        try:
            emp_id = int(v)
        except (TypeError, ValueError):
            continue
        out[emp_id] = r
    return out


def normalise(value):
    """Tidy a value before writing it to the target workbook.

    - ``None`` and blank strings become ``None`` (cell stays empty).
    - Strings are stripped.
    - Anything else is returned unchanged (numbers, datetimes, ...).
    """
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        return v if v else None
    return value


def autofit_column_widths(sheet, min_width: float = 3.0, max_width: float = 40.0) -> None:
    """Set each column width to roughly fit its visible cell contents.

    openpyxl cannot ask Excel for true auto-fit, so this uses a character-
    length heuristic (plus a small padding) and clamps to ``[min_width,
    max_width]``.
    """
    last_row = effective_last_employee_row(sheet, TARGET_FIRST_DATA_ROW)
    last_col = max(sheet.max_column, TARGET_TOTAL_COL)

    for col in range(1, last_col + 1):
        longest = 0
        for row in range(1, last_row + 1):
            value = sheet.cell(row, col).value
            if value is None:
                continue
            if isinstance(value, ArrayFormula) or (
                isinstance(value, str) and value.startswith("=")
            ):
                # Formulas are not what the user sees; prefer nearby header text.
                continue
            if isinstance(value, datetime):
                text = value.strftime("%d-%b-%y")
            elif isinstance(value, date):
                text = value.strftime("%d-%b-%y")
            elif hasattr(value, "hour") and hasattr(value, "minute") and not isinstance(value, datetime):
                # datetime.time cached formula results, etc.
                text = f"{value.hour}:{value.minute:02d}:00"
            else:
                text = str(value)
            if not text:
                continue
            for line in text.splitlines() or [text]:
                longest = max(longest, len(line))
        if longest == 0:
            # Fall back to header cells (rows 1-4) when the column is formula-only.
            for row in range(1, TARGET_FIRST_DATA_ROW):
                value = sheet.cell(row, col).value
                if value is None or isinstance(value, ArrayFormula):
                    continue
                if isinstance(value, (datetime, date)):
                    text = value.strftime("%d-%b-%y")
                elif isinstance(value, str) and value.startswith("="):
                    continue
                else:
                    text = str(value)
                longest = max(longest, len(text))
        if longest == 0:
            continue
        # Excel width units ≈ character count of the default font; pad a bit.
        width = min(max(longest + 2, min_width), max_width)
        sheet.column_dimensions[get_column_letter(col)].width = width


def complete_final(source_path: str, target_path: str) -> None:
    print(f"Loading source: {source_path}")
    src_wb = load_workbook(source_path, data_only=True)
    if SOURCE_SHEET not in src_wb.sheetnames:
        raise SystemExit(f"Source sheet '{SOURCE_SHEET}' not found in {source_path}")
    src = src_wb[SOURCE_SHEET]

    print(f"Loading target: {target_path}")
    tgt_wb = load_workbook(target_path)
    if TARGET_SHEET in tgt_wb.sheetnames:
        tgt = tgt_wb[TARGET_SHEET]
    else:
        tgt = tgt_wb.active
        print(f"  Note: sheet '{TARGET_SHEET}' not found, using active sheet "
              f"'{tgt.title}' instead.")

    merged_map, single_map = build_source_date_maps(src)
    src_id_rows = build_source_id_row_map(src)
    print(f"  Source: {len(src_id_rows)} employees, "
          f"{len(merged_map)} merged-date columns, "
          f"{len(single_map)} single-cell date columns.")

    tgt_date_cols = build_target_date_col_map(tgt)
    tgt_id_rows = build_target_id_row_map(tgt)
    print(f"  Target: {len(tgt_id_rows)} employees, "
          f"{len(tgt_date_cols)} date columns.")

    filled = 0
    blank = 0
    missing_ids: set[int] = set()
    missing_dates: set[date] = set()
    written_cells: list[tuple[int, int]] = []

    for emp_id, tgt_row in tgt_id_rows.items():
        src_row = src_id_rows.get(emp_id)
        if src_row is None:
            missing_ids.add(emp_id)
            continue

        for d, tgt_col in tgt_date_cols.items():
            if d in merged_map:
                src_col = merged_map[d] + 3   # Shortage column
            elif d in single_map:
                src_col = single_map[d]
            else:
                missing_dates.add(d)
                continue

            value = normalise(src.cell(src_row, src_col).value)
            tgt.cell(tgt_row, tgt_col).value = value
            written_cells.append((tgt_row, tgt_col))
            if value is None:
                blank += 1
            else:
                filled += 1

    abbreviated = 0
    reformatted = 0
    for r, c in written_cells:
        cell = tgt.cell(r, c)
        new_val = abbreviate(cell.value)
        if new_val != cell.value:
            cell.value = new_val
            abbreviated += 1
        formatted = format_duration(cell.value)
        if formatted != cell.value:
            cell.value = formatted
            reformatted += 1

    # Column AN ('Total') must keep the template's array formula
    # (=BYROW sum over the date columns). Do NOT overwrite it with static
    # totals — that was turning the formula into plain text like '0:16:00'.
    an_cell = tgt.cell(TARGET_FIRST_DATA_ROW, TARGET_TOTAL_COL)
    if isinstance(an_cell.value, ArrayFormula) or (
        isinstance(an_cell.value, str) and an_cell.value.startswith("=")
    ):
        print(
            f"  Preserved Total formula in column "
            f"{get_column_letter(TARGET_TOTAL_COL)} "
            f"(row {TARGET_FIRST_DATA_ROW})."
        )
    else:
        print(
            f"  Warning: no Total formula found in "
            f"{get_column_letter(TARGET_TOTAL_COL)}{TARGET_FIRST_DATA_ROW}; "
            f"left cell as-is (value={an_cell.value!r})."
        )

    print("  Auto-fitting column widths to contents...")
    autofit_column_widths(tgt)

    tgt_wb.save(target_path)
    print(f"\nDone. {filled} cells written ({blank} cleared), "
          f"{abbreviated} abbreviated, {reformatted} time-formatted "
          f"-> {target_path}")
    if missing_ids:
        preview = ", ".join(str(x) for x in sorted(missing_ids)[:10])
        more = "" if len(missing_ids) <= 10 else f" (+{len(missing_ids)-10} more)"
        print(f"  {len(missing_ids)} target employee ID(s) not found in source: "
              f"{preview}{more}")
    if missing_dates:
        preview = ", ".join(d.isoformat() for d in sorted(missing_dates)[:10])
        more = "" if len(missing_dates) <= 10 else f" (+{len(missing_dates)-10} more)"
        print(f"  {len(missing_dates)} target date(s) not found in source: "
              f"{preview}{more}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Copy attendance values from the Nagwa source into the "
                    "Final Nagwa monthly report.",
    )
    parser.add_argument(
        "-s", "--source",
        default=SOURCE_PATH_DEFAULT,
        help="Source workbook (default: output/Nagwa Technologies.xlsx).",
    )
    parser.add_argument(
        "-t", "--target",
        default=TARGET_PATH_DEFAULT,
        help="Target workbook (default: output/Final Nagwa Technologies.xlsx).",
    )
    args = parser.parse_args()

    if not os.path.isfile(args.source):
        parser.error(f"Source file not found: {args.source}")
    if not os.path.isfile(args.target):
        parser.error(f"Target file not found: {args.target}")

    complete_final(args.source, args.target)


if __name__ == "__main__":
    main()
