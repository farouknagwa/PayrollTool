#!/usr/bin/env python3
"""Replace the calendar columns in Nagwa Technologies.xlsx with the
attendance period covering the 21st of one month through the 20th of the
next month.

The source workbook lives in ``templates/`` and the result is written to
``output/`` using the same file name.

The period is auto-detected from the raw attendance report
(``raw data/Attendance Report.xls``): the first date in the "Attendance
Day" column decides which 21st->20th period to build. For example, a first
date of 28-04-2026 yields the period 21 April -> 20 May.

Mirrors the existing workbook layout:
  - Sun–Thu workdays: 4 columns (in / out / Leave / Shortage), date merged
    across the four columns on row 2, dark-blue header / tan shortage / light
    blue borders.
  - Fri–Sat: single column with the date merged across rows 2–3 (salmon).

Existing date columns from column J onward are wiped (values + merges) and
re-built from scratch using styles captured from the prototype columns
(J:M for weekdays, R for the weekend) BEFORE clearing.
"""

from __future__ import annotations

import argparse
import os
from copy import copy
from datetime import date, datetime, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Paths are resolved relative to the project root (the parent of scripts/).
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
DEFAULT_INPUT = os.path.join(PROJECT_ROOT, "templates", "Nagwa Technologies.xlsx")
DEFAULT_OUTPUT_DIR = os.path.join(PROJECT_ROOT, "output")
DEFAULT_REPORT = os.path.join(PROJECT_ROOT, "raw data", "Attendance Report.xls")

# Columns in the shipped template used as style prototypes (1-based).
DATE_COL_START = 10  # J — first date column / first column of weekday block
WEEKEND_PROTO_OFFSET = 8  # 8 columns past J (= R) is the weekend prototype
WEEKEND_PROTO_COL = DATE_COL_START + WEEKEND_PROTO_OFFSET  # R = 18


# ----------------------------------------------------------------------
# Period detection
# ----------------------------------------------------------------------

def compute_period(ref: date) -> tuple[date, date]:
    """Return the (start, end) of the HR period containing ``ref``.

    Periods run from the 21st of one month through the 20th of the next
    month, inclusive. The period chosen is the one that contains ``ref``:
      - day >= 21 -> period starts on the 21st of ``ref``'s month;
      - day <= 20 -> period starts on the 21st of the previous month.
    """
    if ref.day >= 21:
        start = date(ref.year, ref.month, 21)
    elif ref.month == 1:
        start = date(ref.year - 1, 12, 21)
    else:
        start = date(ref.year, ref.month - 1, 21)

    if start.month == 12:
        end = date(start.year + 1, 1, 20)
    else:
        end = date(start.year, start.month + 1, 20)
    return start, end


def detect_first_attendance_date(report_path: str) -> date:
    """Read the first date in the "Attendance Day" column of the report."""
    df = pd.read_excel(report_path, header=0)
    col = "Attendance Day"
    if col not in df.columns:
        raise SystemExit(
            f"Column {col!r} not found in {report_path!r}. "
            f"Available columns: {list(df.columns)}"
        )
    series = df[col].dropna()
    if series.empty:
        raise SystemExit(f"No dates found in column {col!r} of {report_path!r}.")

    first = series.iloc[0]
    if isinstance(first, str):
        return datetime.strptime(first.strip(), "%d/%m/%Y").date()
    return pd.to_datetime(first).date()


def daterange(start: date, end: date):
    cur = start
    while cur <= end:
        yield cur
        cur += timedelta(days=1)


def is_egypt_weekend(d: date) -> bool:
    """Friday = 4, Saturday = 5 in datetime.weekday() (Monday = 0)."""
    return d.weekday() in (4, 5)


# ----------------------------------------------------------------------
# Style capture / apply
# ----------------------------------------------------------------------

def capture_column_styles(sheet, col: int, max_row: int) -> list[dict]:
    """Snapshot the per-row styles of ``col`` from row 1 through ``max_row``."""
    snapshots: list[dict] = []
    for r in range(1, max_row + 1):
        c = sheet.cell(r, col)
        snapshots.append(
            {
                "has_style": c.has_style,
                "font": copy(c.font),
                "fill": copy(c.fill),
                "border": copy(c.border),
                "alignment": copy(c.alignment),
                "number_format": c.number_format,
            }
        )
    return snapshots


def apply_styles_to_column(sheet, dest_col: int, styles: list[dict]) -> None:
    for r, s in enumerate(styles, start=1):
        c = sheet.cell(r, dest_col)
        if s["has_style"]:
            c.font = s["font"]
            c.fill = s["fill"]
            c.border = s["border"]
            c.alignment = s["alignment"]
        c.number_format = s["number_format"]


# ----------------------------------------------------------------------
# Block writers
# ----------------------------------------------------------------------

def apply_weekday_block(
    sheet,
    day: date,
    start_col: int,
    max_row: int,
    row1_values: list[int],
    weekday_proto: list[list[dict]],
    weekday_widths: list[float | None],
) -> None:
    """Write a 4-column workday block at ``start_col``."""
    for off in range(4):
        apply_styles_to_column(sheet, start_col + off, weekday_proto[off])

    for i, v in enumerate(row1_values):
        sheet.cell(1, start_col + i).value = v

    dt = datetime(day.year, day.month, day.day)
    sheet.merge_cells(
        start_row=2,
        start_column=start_col,
        end_row=2,
        end_column=start_col + 3,
    )
    sheet.cell(2, start_col).value = dt

    for i, lab in enumerate(("in", "out", "Leave", "Shortage")):
        sheet.cell(3, start_col + i).value = lab

    for r in range(4, max_row + 1):
        for c in range(start_col, start_col + 4):
            sheet.cell(r, c).value = None

    for off in range(4):
        w = weekday_widths[off]
        if w is not None:
            sheet.column_dimensions[get_column_letter(start_col + off)].width = w


def apply_weekend_block(
    sheet,
    day: date,
    col: int,
    max_row: int,
    row1_val: int,
    weekend_proto: list[dict],
    weekend_width: float | None,
) -> None:
    """Write a single weekend column at ``col``."""
    apply_styles_to_column(sheet, col, weekend_proto)

    sheet.cell(1, col).value = row1_val

    dt = datetime(day.year, day.month, day.day)
    sheet.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
    sheet.cell(2, col).value = dt

    for r in range(4, max_row + 1):
        sheet.cell(r, col).value = None

    if weekend_width is not None:
        sheet.column_dimensions[get_column_letter(col)].width = weekend_width


# ----------------------------------------------------------------------
# Sheet wipe
# ----------------------------------------------------------------------

def clear_date_area(sheet, max_row: int) -> int:
    """Unmerge and clear values in the date column area (col J onward).

    Returns the highest column index that contained data.
    """
    last_col = max(sheet.max_column, DATE_COL_START - 1)

    # Unmerge any merged ranges that fall entirely inside the date area.
    to_unmerge = [
        str(mc)
        for mc in list(sheet.merged_cells.ranges)
        if mc.min_col >= DATE_COL_START
    ]
    for r in to_unmerge:
        sheet.unmerge_cells(r)

    # Clear values in every cell from col J to the last used column.
    for c in range(DATE_COL_START, last_col + 1):
        for r in range(1, max_row + 1):
            sheet.cell(r, c).value = None

    return last_col


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------

def replace_calendar_period(
    path: str,
    out_path: str,
    ref_date: date,
) -> None:
    start_date, end_date = compute_period(ref_date)

    wb = load_workbook(path)
    if "Nagwa Technologies" not in wb.sheetnames:
        raise SystemExit(
            f"Sheet 'Nagwa Technologies' not found in {path!r}. "
            f"Available sheets: {wb.sheetnames}"
        )
    sheet = wb["Nagwa Technologies"]
    max_row = sheet.max_row

    # 1. Capture prototype styles from the existing template BEFORE clearing.
    weekday_proto = [
        capture_column_styles(sheet, DATE_COL_START + i, max_row) for i in range(4)
    ]
    weekend_proto = capture_column_styles(sheet, WEEKEND_PROTO_COL, max_row)

    weekday_widths = [
        sheet.column_dimensions[get_column_letter(DATE_COL_START + i)].width
        for i in range(4)
    ]
    weekend_width = sheet.column_dimensions[get_column_letter(WEEKEND_PROTO_COL)].width

    # 2. Capture the existing row-1 counter at column J (so the rebuilt
    #    period continues using the same numbering scheme).
    j1_value = sheet.cell(1, DATE_COL_START).value
    if isinstance(j1_value, (int, float)):
        counter = int(j1_value)
    else:
        last_left = sheet.cell(1, DATE_COL_START - 1).value
        counter = int(last_left) + 1 if isinstance(last_left, (int, float)) else 1

    # 3. Wipe existing date columns (values + merges).
    last_col = clear_date_area(sheet, max_row)

    # 4. Build the new period from column J.
    col_cursor = DATE_COL_START
    weekday_count = 0
    weekend_count = 0
    for d in daterange(start_date, end_date):
        if is_egypt_weekend(d):
            apply_weekend_block(
                sheet, d, col_cursor, max_row, counter, weekend_proto, weekend_width
            )
            counter += 1
            col_cursor += 1
            weekend_count += 1
        else:
            nums = [counter, counter + 1, counter + 2, counter + 3]
            apply_weekday_block(
                sheet,
                d,
                col_cursor,
                max_row,
                nums,
                weekday_proto,
                weekday_widths,
            )
            counter += 4
            col_cursor += 4
            weekday_count += 1

    # 5. Drop column-dimension entries for any leftover columns past the
    #    new period so that extra-wide stale columns from the previous
    #    period don't linger.
    for c in range(col_cursor, last_col + 1):
        sheet.column_dimensions.pop(get_column_letter(c), None)

    wb.save(out_path)
    wb.close()

    new_cols = col_cursor - DATE_COL_START
    print(
        f"Replaced calendar with period {start_date.isoformat()} "
        f"-> {end_date.isoformat()} "
        f"({weekday_count} workday block(s), {weekend_count} weekend day(s), "
        f"{new_cols} column(s) starting at {get_column_letter(DATE_COL_START)})."
    )
    print(f"Saved: {out_path}")


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "-i",
        "--input",
        default=DEFAULT_INPUT,
        help="Source workbook (default: templates/Nagwa Technologies.xlsx)",
    )
    p.add_argument(
        "-o",
        "--output",
        default=None,
        help=(
            "Output workbook. Defaults to the output/ folder using the "
            "same file name as the input."
        ),
    )
    p.add_argument(
        "-r",
        "--report",
        default=DEFAULT_REPORT,
        help=(
            "Attendance report used to detect the period "
            "(default: raw data/Attendance Report.xls)."
        ),
    )
    args = p.parse_args()

    inp = os.path.abspath(args.input)
    if not os.path.isfile(inp):
        raise SystemExit(f"Input not found: {inp}")

    report = os.path.abspath(args.report)
    if not os.path.isfile(report):
        raise SystemExit(f"Attendance report not found: {report}")

    first_date = detect_first_attendance_date(report)

    if args.output:
        out = os.path.abspath(args.output)
    else:
        out = os.path.join(DEFAULT_OUTPUT_DIR, os.path.basename(inp))

    os.makedirs(os.path.dirname(out), exist_ok=True)

    print(f"Detected first attendance date: {first_date.isoformat()}")
    replace_calendar_period(inp, out, first_date)


if __name__ == "__main__":
    main()
