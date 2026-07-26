import os
import re
from datetime import datetime, time, timedelta

import pandas as pd
from openpyxl import load_workbook


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)

RAW_DATA_DIR = os.path.join(PROJECT_ROOT, "raw data")

# Report base names (without extension). Each report is resolved at read
# time and accepted with either an .xls or .xlsx extension.
ATTENDANCE_BASENAME = "Attendance Report"
VACATIONS_BASENAME = "Employee Transactions_vacations"
RESIGNATIONS_BASENAME = "Resignations"
PUBLIC_HOLIDAY_BASENAME = "Public Holiday"
ABSENCES_BASENAME = "Absence Report"
PERMISSIONS_BASENAME = "Nagwa_Permission_Request_permission_details"
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "output")
NAGWA_PATH = os.path.join(OUTPUT_DIR, "Nagwa Technologies.xlsx")


def resolve_report_path(base_name, folder=RAW_DATA_DIR):
    """Return the path to ``<folder>/<base_name>.<ext>`` for whichever of
    ``.xlsx`` or ``.xls`` actually exists on disk.

    Raises ``FileNotFoundError`` if neither extension is present.
    """
    for ext in ("xlsx", "xls"):
        path = os.path.join(folder, f"{base_name}.{ext}")
        if os.path.exists(path):
            return path
    raise FileNotFoundError(
        f"Report '{base_name}' not found in '{folder}/' "
        f"(tried .xlsx and .xls)."
    )


def _looks_like_employee_code(value):
    """True if a cell value looks like a numeric employee code (a data row)."""
    if value is None:
        return False
    s = str(value).strip()
    if s in ("", "nan", "None"):
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def _read_positional_report(base_name, folder=RAW_DATA_DIR, code_col=0):
    """Read a report whose columns are addressed positionally.

    These raw exports carry a single header row, but the reader historically
    assumed two and used ``skiprows=2``, which silently discarded the first
    data record (e.g. the first employee's absence). Instead of hard-coding a
    skip count, detect the first data row as the first row whose ``code_col``
    cell holds a numeric employee code and drop everything above it. This
    tolerates reports with one or two header/pre-amble rows without losing the
    first record.
    """
    raw = pd.read_excel(
        resolve_report_path(base_name, folder),
        engine="calamine",
        header=None,
    )
    start = 0
    for i in range(len(raw)):
        if _looks_like_employee_code(raw.iat[i, code_col]):
            start = i
            break
    return raw.iloc[start:].reset_index(drop=True)

WORKDAY_START = time(8, 0)
WORKDAY_END_NORMAL = time(16, 0)
WORKDAY_END_RAMADAN = time(14, 30)

RAMADAN_START = datetime(2026, 2, 20).date()
RAMADAN_END = datetime(2026, 3, 18).date()
FULL_DAY_NORMAL = 8 * 60        # 480 minutes
FULL_DAY_RAMADAN = 6 * 60 + 30  # 390 minutes

PERMITTED_WINDOW_END_NORMAL = time(18, 0)    # 6:00 PM
PERMITTED_WINDOW_END_RAMADAN = time(16, 30)  # 4:30 PM

# --- Schedule-based permitted-window ends (non-Ramadan, non-special-rule) ---
# Column E ("Schedule") in the Nagwa sheet drives the permitted attendance
# window for each employee. The full day is always 8 hours; only the window
# end (and therefore the latest minute that counts toward presence) changes.
#   Undefined / Standard -> 8:00 AM - 4:00 PM
#   Flexy                -> 8:00 AM - 6:00 PM
#   Flexy9               -> 8:00 AM - 5:00 PM
# Anything else falls back to "Standard" (8:00 AM - 4:00 PM).
SCHEDULE_WINDOW_END = {
    "standard":  time(16, 0),
    "undefined": time(16, 0),
    "flexy":     time(18, 0),
    "flexy9":    time(17, 0),
}
DEFAULT_SCHEDULE_WINDOW_END = time(16, 0)

# --- Special restricted-window rule for two specific employees ---
# Employees 1052 and 100 alternate, each "period" running from the 21st of
# one month to the 20th of the next.  The employee whose period it is gets a
# restricted non-Ramadan window of 8:00 AM - 4:00 PM (instead of 6:00 PM);
# the other employee keeps the normal window.
# Period 0 (21 Feb 2026 - 20 Mar 2026): employee 1052 is restricted.
# Period 1 (21 Mar 2026 - 20 Apr 2026): employee 100  is restricted.
# ... and so on, alternating each period.
SPECIAL_RULE_EMPLOYEE_A = 1052
SPECIAL_RULE_EMPLOYEE_B = 100
SPECIAL_RULE_ANCHOR = datetime(2026, 2, 21).date()
SPECIAL_RULE_PAIRS = [(SPECIAL_RULE_EMPLOYEE_A, SPECIAL_RULE_EMPLOYEE_B, SPECIAL_RULE_ANCHOR)]
SPECIAL_RULE_EMPLOYEES = {SPECIAL_RULE_EMPLOYEE_A, SPECIAL_RULE_EMPLOYEE_B}
PERMITTED_WINDOW_END_RESTRICTED = time(16, 0)  # 4:00 PM

# --- Work-Mission lunch-gap exemption ---
# When an employee has at least one Work Mission entry on a given day, any
# uncovered minutes that fall inside the lunch window are not counted as
# shortage. The lunch window itself shifts on 26 Apr 2026:
#   * Before 26 Apr 2026:        12:00 PM – 1:00 PM
#   * On / after 26 Apr 2026:    12:30 PM – 1:30 PM
LUNCH_WINDOW_SWITCH_DATE = datetime(2026, 4, 26).date()
LUNCH_WINDOW_BEFORE = (time(12, 0), time(13, 0))
LUNCH_WINDOW_FROM = (time(12, 30), time(13, 30))


def _lunch_window_for(att_date):
    """Return the (start, end) lunch window applicable on ``att_date``."""
    if att_date is None or att_date < LUNCH_WINDOW_SWITCH_DATE:
        return LUNCH_WINDOW_BEFORE
    return LUNCH_WINDOW_FROM


# --- Per-employee 1-hour shortage reduction windows ---
# For each listed employee, every daily shortage that falls within the given
# (inclusive) date range is reduced by 60 minutes. Results are clamped at zero
# so a shortage of less than one hour becomes 0:00. A None bound means the
# range is open on that end.
HOUR_REDUCTION_WINDOWS = {
    514:  (datetime(2025, 3, 9).date(),   datetime(2026, 4, 7).date()),
    2148: (datetime(2024, 9, 18).date(),  datetime(2026, 6, 20).date()),
    350:  (datetime(2025, 8, 3).date(),   datetime(2026, 10, 27).date()),
    1809: (None,                          datetime(2026, 11, 1).date()),
    822:  (datetime(2025, 3, 10).date(),  datetime(2026, 12, 10).date()),
    315:  (datetime(2025, 9, 7).date(),   datetime(2027, 6, 6).date()),
}


def convert_date(date_str):
    """Convert DD/MM/YYYY to DD-Mon-YY format."""
    try:
        dt = datetime.strptime(date_str, "%d/%m/%Y")
        return dt.strftime("%d-%b-%y")
    except ValueError:
        return date_str


def parse_duration(time_str):
    """Parse H:MM or HH:MM duration string into total minutes."""
    if not time_str or str(time_str).strip() in ("", "nan"):
        return 0
    parts = str(time_str).strip().split(":")
    if len(parts) == 2:
        return int(parts[0]) * 60 + int(parts[1])
    return 0


def minutes_to_hhmm(total_minutes):
    """Convert total minutes back to H:MM string."""
    h = total_minutes // 60
    m = total_minutes % 60
    return f"{h}:{m:02d}"


def _special_rule_period_index(att_date, anchor=None):
    """Return the alternating-period index for the special rule, or -1 if the
    date is before the configured anchor.

    Each period runs from the 21st of one month to the 20th of the next, so a
    date with day < 21 still belongs to the period that started on the 21st of
    the previous month.
    """
    anchor = anchor or SPECIAL_RULE_ANCHOR
    if att_date < anchor:
        return -1
    if att_date.day >= 21:
        period_month, period_year = att_date.month, att_date.year
    elif att_date.month == 1:
        period_month, period_year = 12, att_date.year - 1
    else:
        period_month, period_year = att_date.month - 1, att_date.year
    return (period_year - anchor.year) * 12 + (period_month - anchor.month)


def has_restricted_window(emp_code, att_date):
    """True iff the 8 AM - 4 PM restricted non-Ramadan window applies to
    ``emp_code`` on ``att_date``.

    For each configured pair, even periods restrict Employee A; odd periods
    restrict Employee B. Returns False outside the rule's scope.
    """
    if emp_code is None:
        return False
    for employee_a, employee_b, anchor in SPECIAL_RULE_PAIRS:
        if emp_code not in (employee_a, employee_b):
            continue
        period_index = _special_rule_period_index(att_date, anchor)
        if period_index < 0:
            continue
        restricted_emp = employee_a if period_index % 2 == 0 else employee_b
        if emp_code == restricted_emp:
            return True
    return False


def _permitted_window_end(att_date, emp_code, schedule):
    """Return the permitted-window end time for a given employee/date/schedule.

    Encapsulates the Ramadan / restricted / special-rule / schedule logic so
    both ``calculate_shortage`` and the leave-rule overrides stay in sync.
    """
    if att_date and RAMADAN_START <= att_date <= RAMADAN_END:
        return PERMITTED_WINDOW_END_RAMADAN
    if att_date is not None and has_restricted_window(emp_code, att_date):
        return PERMITTED_WINDOW_END_RESTRICTED
    if emp_code in SPECIAL_RULE_EMPLOYEES:
        # Special-rule employees keep their original 8 AM - 6 PM window
        # on non-restricted days, ignoring the Schedule column.
        return PERMITTED_WINDOW_END_NORMAL
    sched_key = (schedule or "").strip().lower()
    return SCHEDULE_WINDOW_END.get(sched_key, DEFAULT_SCHEDULE_WINDOW_END)


def calculate_shortage(late_str, entry_time_str, exit_time_str, att_date=None, emp_code=None, schedule=None):
    """
    Shortage based on effective presence within the permitted attendance window.

    Permitted window:
      - Ramadan (20 Feb 2026 – 18 Mar 2026): 8:00 AM – 4:30 PM, full day = 6:30
      - Regular days, by Schedule (column E in the Nagwa sheet):
          * Undefined / Standard -> 8:00 AM – 4:00 PM
          * Flexy                -> 8:00 AM – 6:00 PM
          * Flexy9               -> 8:00 AM – 5:00 PM
        Full day is 8:00 in every case.
      - Special rule (employees 1052 / 100, alternating monthly from
        21 Feb 2026): on non-Ramadan days the window is 8:00 AM – 4:00 PM,
        regardless of their Schedule value.

    Only the overlap between (entry, exit) and the window counts.
    If effective presence >= full_day -> shortage = late
    If effective presence <  full_day -> shortage = full_day - effective presence
    """
    late_minutes = parse_duration(late_str)

    if att_date and RAMADAN_START <= att_date <= RAMADAN_END:
        full_day = FULL_DAY_RAMADAN
    else:
        full_day = FULL_DAY_NORMAL
    window_end = _permitted_window_end(att_date, emp_code, schedule)

    in_time = parse_time_value(entry_time_str)
    out_time = parse_time_value(exit_time_str)

    if in_time is not None and out_time is not None:
        lo = max(_time_to_min(in_time), _time_to_min(WORKDAY_START))
        hi = min(_time_to_min(out_time), _time_to_min(window_end))
        effective_minutes = max(0, hi - lo)
    else:
        effective_minutes = 0

    if effective_minutes >= full_day:
        return minutes_to_hhmm(late_minutes)
    else:
        shortage_minutes = full_day - effective_minutes
        return minutes_to_hhmm(shortage_minutes)


def extract_date_from_cell(cell_value):
    """Extract a date object from a Nagwa header cell (datetime or string like '29/01/2026 PUBLIC Holiday')."""
    if isinstance(cell_value, datetime):
        return cell_value.date()
    if isinstance(cell_value, str):
        match = re.match(r"(\d{2}/\d{2}/\d{4})", cell_value)
        if match:
            return datetime.strptime(match.group(1), "%d/%m/%Y").date()
    return None


def build_date_column_map(sheet):
    """
    Build a dict: date -> column index of the 'in' sub-column.
    Only includes dates that have the 4 sub-columns (in/out/Leave/Shortage).
    """
    date_col_map = {}
    for mc in sheet.merged_cells.ranges:
        if mc.min_row != 2 or mc.min_col < 10:
            continue
        span = mc.max_col - mc.min_col + 1
        if span != 4:
            continue
        sub_header = sheet.cell(3, mc.min_col).value
        if sub_header and sub_header.lower() == "in":
            cell_value = sheet.cell(2, mc.min_col).value
            dt = extract_date_from_cell(cell_value)
            if dt:
                date_col_map[dt] = mc.min_col
    return date_col_map


def build_single_date_column_map(sheet, date_col_map):
    """
    Build a dict: date -> column index for dates whose header in row 2 has
    no in/out/Leave/Shortage sub-columns (i.e. a single cell under the date).
    """
    merged_info = {}
    for mc in sheet.merged_cells.ranges:
        if mc.min_row <= 2 <= mc.max_row:
            for c in range(mc.min_col, mc.max_col + 1):
                merged_info[c] = mc.min_col

    single_map = {}
    for col in range(10, sheet.max_column + 1):
        cell_value = sheet.cell(2, col).value
        if cell_value is None:
            continue
        dt = extract_date_from_cell(cell_value)
        if dt is None or dt in date_col_map:
            continue
        if col in merged_info and merged_info[col] != col:
            continue
        single_map[dt] = col
    return single_map


def effective_last_employee_row(sheet, first_row=4, id_col=1):
    """Return the last employee row, ignoring formatting-only trailing rows."""
    last = first_row - 1
    blank_run = 0
    for row in range(first_row, sheet.max_row + 1):
        if sheet.cell(row, id_col).value is None:
            if last >= first_row:
                blank_run += 1
                if blank_run >= 50:
                    break
            continue
        last = row
        blank_run = 0
    return max(last, first_row - 1)


def build_code_row_map(sheet):
    """Build a dict: employee_code (int) -> row number in the Nagwa sheet."""
    code_row_map = {}
    last_row = effective_last_employee_row(sheet)
    for r in range(4, last_row + 1):
        code = sheet.cell(r, 1).value
        if code is not None:
            code_row_map[int(code)] = r
    return code_row_map


def build_code_schedule_map(sheet, code_row_map):
    """Build a dict: employee_code (int) -> normalized schedule string.

    The Schedule lives in column E (5) of the Nagwa sheet. Values are
    lower-cased and stripped so callers can do a direct dict lookup against
    ``SCHEDULE_WINDOW_END``. Empty/missing cells map to 'standard'.
    """
    schedule_map = {}
    for emp_code, row in code_row_map.items():
        raw = sheet.cell(row, 5).value
        if raw is None:
            schedule_map[emp_code] = "standard"
            continue
        normalized = str(raw).strip().lower()
        schedule_map[emp_code] = normalized if normalized else "standard"
    return schedule_map


def build_code_status_map(sheet, code_row_map):
    """Build a dict: employee_code (int) -> normalized employment-status string.

    Employment Status lives in column C (3) of the Nagwa sheet. Values are
    lower-cased and stripped; empty/missing cells map to ''.
    """
    status_map = {}
    for emp_code, row in code_row_map.items():
        raw = sheet.cell(row, 3).value
        if raw is None:
            status_map[emp_code] = ""
            continue
        status_map[emp_code] = str(raw).strip().lower()
    return status_map


def is_absence_exempt(emp_code, schedule_map, status_map):
    """True if this employee must never be marked absent.

    HR rule: Schedule = Undefined, or Employment Status = Challenged 5%,
    are not eligible for absence calculations.
    """
    if (schedule_map.get(emp_code) or "").strip().lower() == "undefined":
        return True
    status = (status_map.get(emp_code) or "").strip().lower()
    return "challenged" in status and "5%" in status


def _coerce_date(value):
    """Best-effort conversion of a cell value into a ``date`` (or None)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    # pandas Timestamp is a datetime subclass, so the check above covers it.
    s = str(value).strip()
    if s in ("", "nan", "None"):
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%b-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _dates_from_ddmmyyyy(series):
    """Return the list of ``date`` values parsed from a DD/MM/YYYY series."""
    parsed = pd.to_datetime(series, format="%d/%m/%Y", errors="coerce").dropna()
    return [ts.date() for ts in parsed]


def compute_coverage_window():
    """Return (min_date, max_date) actually covered by the source reports.

    The HR period built into the sheet can extend past the data that has been
    reported so far (e.g. a 21st->20th period while the reports only cover up
    to the run date). The closed-world absence sweep must never infer an
    absence for a day outside this window, since no data exists for it yet.
    Dates are drawn from the Attendance and Absence reports; either may be
    missing. Returns ``(None, None)`` when no dated rows are available.
    """
    dates = []
    try:
        dates += _dates_from_ddmmyyyy(read_attendance()["Attendance Day"])
    except (FileNotFoundError, KeyError):
        pass
    try:
        dates += _dates_from_ddmmyyyy(read_absences()["Absence Date"])
    except (FileNotFoundError, KeyError):
        pass
    if not dates:
        return None, None
    return min(dates), max(dates)


def build_code_employment_map(sheet, code_row_map):
    """Build a dict: employee_code -> (hire_date, termination_date).

    Hire Date lives in column G (7) and Termination Date in column H (8) of the
    Nagwa sheet. Either bound is ``None`` when the cell is empty/unparseable,
    meaning "open on that end" for employment-window checks.
    """
    employment_map = {}
    for emp_code, row in code_row_map.items():
        hire = _coerce_date(sheet.cell(row, 7).value)
        term = _coerce_date(sheet.cell(row, 8).value)
        employment_map[emp_code] = (hire, term)
    return employment_map


def read_attendance():
    """Read the Attendance Report and return a cleaned DataFrame."""
    df = _read_positional_report(ATTENDANCE_BASENAME)
    df.columns = [
        "Code", "Employee Name", "Card Id", "Weekday", "Attendance Day",
        "Work Time", "Entry Time", "Late", "TA Permissions", "Early Overtime",
        "End Time", "Exit Time", "Overtime", "Early Out Leave", "On Call",
        "Actual Hours",
    ]
    df = df.dropna(subset=["Code"])
    df["Code"] = df["Code"].astype(str).str.strip()
    return df


def read_absences():
    """Read the Absence Report and return a cleaned DataFrame."""
    df = _read_positional_report(ABSENCES_BASENAME)
    df = df.dropna(subset=[df.columns[0]])
    df.rename(columns={df.columns[0]: "Employee Code", df.columns[4]: "Absence Date"}, inplace=True)
    df["Employee Code"] = df["Employee Code"].astype(str).str.strip()
    return df


def fill_absences(sheet, date_col_map, code_row_map, schedule_map=None, status_map=None):
    """Fill 'absent' in in/out/shortage for each absence record.

    Employees with Schedule = Undefined or Employment Status = Challenged 5%
    are skipped (not eligible for absence calculations).
    """
    print("\nReading Absence Report...")
    absences_df = read_absences()
    print(f"  {len(absences_df)} absence records loaded.")

    filled = 0
    skipped_code = 0
    skipped_date = 0
    skipped_exempt = 0
    schedule_map = schedule_map or {}
    status_map = status_map or {}

    for _, row in absences_df.iterrows():
        try:
            emp_code = int(float(row["Employee Code"]))
        except (ValueError, TypeError):
            skipped_code += 1
            continue

        if emp_code not in code_row_map:
            skipped_code += 1
            continue

        if is_absence_exempt(emp_code, schedule_map, status_map):
            skipped_exempt += 1
            continue

        abs_date_str = str(row["Absence Date"]).strip()
        try:
            abs_date = datetime.strptime(abs_date_str, "%d/%m/%Y").date()
        except ValueError:
            skipped_date += 1
            continue

        if abs_date not in date_col_map:
            skipped_date += 1
            continue

        nagwa_row = code_row_map[emp_code]
        in_col = date_col_map[abs_date]
        out_col = in_col + 1
        shortage_col = in_col + 3

        sheet.cell(nagwa_row, in_col).value = "absent"
        sheet.cell(nagwa_row, out_col).value = "absent"
        sheet.cell(nagwa_row, shortage_col).value = "absent"
        filled += 1

    print(f"\nAbsences done! {filled} day-cells filled.")
    if skipped_code:
        print(f"  {skipped_code} rows skipped (employee code not found in Nagwa sheet).")
    if skipped_date:
        print(f"  {skipped_date} rows skipped (date not found in Nagwa sheet).")
    if skipped_exempt:
        print(
            f"  {skipped_exempt} rows skipped "
            f"(Undefined schedule or Challenged 5% — not eligible for absence)."
        )


def main():
    print("Reading Attendance Report...")
    attendance_df = read_attendance()
    print(f"  {len(attendance_df)} attendance records loaded.")

    print("Loading Nagwa Technologies workbook...")
    wb = load_workbook(NAGWA_PATH)
    sheet = wb["Nagwa Technologies"]

    date_col_map = build_date_column_map(sheet)
    single_date_col_map = build_single_date_column_map(sheet, date_col_map)
    code_row_map = build_code_row_map(sheet)
    code_schedule_map = build_code_schedule_map(sheet, code_row_map)
    code_status_map = build_code_status_map(sheet, code_row_map)
    code_employment_map = build_code_employment_map(sheet, code_row_map)

    print(f"  {len(date_col_map)} workday date columns found.")
    print(f"  {len(single_date_col_map)} single-cell date columns found.")
    print(f"  {len(code_row_map)} employees found in Nagwa sheet.")

    filled = 0
    skipped_code = 0
    skipped_date = 0

    for _, row in attendance_df.iterrows():
        try:
            emp_code = int(float(row["Code"]))
        except (ValueError, TypeError):
            skipped_code += 1
            continue

        if emp_code not in code_row_map:
            skipped_code += 1
            continue

        att_date_str = str(row["Attendance Day"]).strip()
        try:
            att_date = datetime.strptime(att_date_str, "%d/%m/%Y").date()
        except ValueError:
            skipped_date += 1
            continue

        if att_date not in date_col_map:
            skipped_date += 1
            continue

        nagwa_row = code_row_map[emp_code]
        in_col = date_col_map[att_date]
        out_col = in_col + 1
        leave_col = in_col + 2
        shortage_col = in_col + 3

        entry_time = str(row["Entry Time"]).strip() if pd.notna(row["Entry Time"]) else ""
        exit_time = str(row["Exit Time"]).strip() if pd.notna(row["Exit Time"]) else ""
        late_val = str(row["Late"]).strip() if pd.notna(row["Late"]) else "0:00"

        schedule = code_schedule_map.get(emp_code)
        shortage = calculate_shortage(
            late_val, entry_time, exit_time, att_date, emp_code, schedule
        )

        sheet.cell(nagwa_row, in_col).value = entry_time
        sheet.cell(nagwa_row, out_col).value = exit_time
        # Leave column stays empty
        sheet.cell(nagwa_row, shortage_col).value = shortage

        filled += 1

    print(f"\nDone! {filled} cells filled.")
    if skipped_code:
        print(f"  {skipped_code} rows skipped (employee code not found in Nagwa sheet).")
    if skipped_date:
        print(f"  {skipped_date} rows skipped (date not found in Nagwa sheet).")

    # --- Absences ---
    fill_absences(
        sheet, date_col_map, code_row_map, code_schedule_map, code_status_map
    )

    # --- Vacations ---
    fill_vacations(sheet, date_col_map, code_row_map, single_date_col_map)

    # --- Resignations ---
    fill_resignations(sheet, date_col_map, code_row_map)

    # --- Public Holidays ---
    fill_public_holidays(sheet, date_col_map, code_row_map)

    # --- Permissions ---
    fill_permissions(sheet, date_col_map, code_row_map)

    # --- Shortage recalculation based on leave rules ---
    recalculate_shortage_from_leave(sheet, date_col_map, code_row_map, code_schedule_map)

    # --- Permitted Delays deduction ---
    apply_permitted_delays(sheet, date_col_map, code_row_map)

    # --- Work-Mission lunch-gap exemption ---
    apply_work_mission_lunch_exemption(sheet, date_col_map, code_row_map)

    # --- Per-employee 1-hour shortage reduction (specific date windows) ---
    apply_hour_reduction(sheet, date_col_map, code_row_map)

    # --- Missing Punch sweep (final step) ---
    fill_missing_punches(
        sheet, date_col_map, code_row_map, code_schedule_map, code_status_map
    )

    # --- WFH / Workday leave overrides ---
    apply_wfh_and_workday_overrides(sheet, date_col_map, code_row_map)

    # --- Full-day absence safety net (closed-world; must run last) ---
    coverage_start, coverage_end = compute_coverage_window()
    fill_full_day_absences(
        sheet, date_col_map, code_row_map, code_employment_map,
        coverage_start, coverage_end,
        code_schedule_map, code_status_map,
    )

    wb.save(NAGWA_PATH)
    print(f"\nSaved to {NAGWA_PATH}")


def read_vacations():
    """Read the Employee Transactions (vacations) report and return a cleaned DataFrame."""
    df = _read_positional_report(VACATIONS_BASENAME)
    df.columns = [
        "Employee Code", "Employee Name", "Date", "Vacation End Date",
        "Vacation Type", "Vacation Days", "Deduction Amount", "Delegate Code",
        "Delegate Name", "Place To Be", "Transaction Status", "Notes",
    ]
    df = df.dropna(subset=["Employee Code"])
    df["Employee Code"] = df["Employee Code"].astype(str).str.strip()
    return df


def fill_vacations(sheet, date_col_map, code_row_map, single_date_col_map=None):
    """Fill vacation type into the in/out columns for each day of each vacation range.

    If a date falls within the vacation range but only has a single cell under
    the date header (no in/out/Leave/Shortage), that single cell is filled too.
    """
    if single_date_col_map is None:
        single_date_col_map = {}
    print("\nReading Vacations Report...")
    vacations_df = read_vacations()
    print(f"  {len(vacations_df)} vacation records loaded.")

    filled = 0
    skipped_code = 0
    skipped_date = 0

    for _, row in vacations_df.iterrows():
        try:
            emp_code = int(float(row["Employee Code"]))
        except (ValueError, TypeError):
            skipped_code += 1
            continue

        if emp_code not in code_row_map:
            skipped_code += 1
            continue

        start_str = str(row["Date"]).strip()
        end_str = str(row["Vacation End Date"]).strip()
        vacation_type = str(row["Vacation Type"]).strip() if pd.notna(row["Vacation Type"]) else ""

        try:
            start_date = datetime.strptime(start_str, "%d/%m/%Y").date()
            end_date = datetime.strptime(end_str, "%d/%m/%Y").date()
        except ValueError:
            skipped_date += 1
            continue

        nagwa_row = code_row_map[emp_code]
        current_date = start_date
        while current_date <= end_date:
            if current_date in date_col_map:
                in_col = date_col_map[current_date]
                out_col = in_col + 1
                shortage_col = in_col + 3
                sheet.cell(nagwa_row, in_col).value = vacation_type
                sheet.cell(nagwa_row, out_col).value = vacation_type
                sheet.cell(nagwa_row, shortage_col).value = vacation_type
                filled += 1
            elif current_date in single_date_col_map:
                col = single_date_col_map[current_date]
                sheet.cell(nagwa_row, col).value = vacation_type
                filled += 1
            else:
                skipped_date += 1
            current_date += timedelta(days=1)

    print(f"\nVacations done! {filled} day-cells filled.")
    if skipped_code:
        print(f"  {skipped_code} rows skipped (employee code not found in Nagwa sheet).")
    if skipped_date:
        print(f"  {skipped_date} day-slots skipped (date not found in Nagwa sheet).")


def read_resignations():
    """Read the Resignations report and return a cleaned DataFrame."""
    df = pd.read_excel(resolve_report_path(RESIGNATIONS_BASENAME), engine="calamine")
    df.columns = df.columns.str.strip()
    df = df.dropna(subset=["ID"])
    df["ID"] = df["ID"].astype(int)
    return df


def fill_resignations(sheet, date_col_map, code_row_map):
    """For each resigned employee, fill 'Resigned' in in/out/shortage from resignation date onward."""
    print("\nReading Resignations Report...")
    try:
        resignations_df = read_resignations()
    except FileNotFoundError:
        print("  Resignations report not found; skipping (optional).")
        return
    print(f"  {len(resignations_df)} resignation records loaded.")

    sorted_dates = sorted(date_col_map.keys())
    filled = 0
    skipped_code = 0

    for _, row in resignations_df.iterrows():
        emp_code = int(row["ID"])
        if emp_code not in code_row_map:
            skipped_code += 1
            continue

        res_date_val = row["Resignation Date"]
        if isinstance(res_date_val, pd.Timestamp):
            res_date = res_date_val.date()
        else:
            res_date = datetime.strptime(str(res_date_val).strip(), "%d/%m/%Y").date()

        nagwa_row = code_row_map[emp_code]
        for d in sorted_dates:
            if d >= res_date:
                in_col = date_col_map[d]
                out_col = in_col + 1
                shortage_col = in_col + 3
                sheet.cell(nagwa_row, in_col).value = "Resigned"
                sheet.cell(nagwa_row, out_col).value = "Resigned"
                sheet.cell(nagwa_row, shortage_col).value = "Resigned"
                filled += 1

    print(f"\nResignations done! {filled} day-cells filled.")
    if skipped_code:
        print(f"  {skipped_code} rows skipped (employee code not found in Nagwa sheet).")


def read_public_holidays():
    """Read the Public Holiday report and return a cleaned DataFrame."""
    df = pd.read_excel(resolve_report_path(PUBLIC_HOLIDAY_BASENAME), engine="calamine")
    df.columns = df.columns.str.strip()
    return df


def fill_public_holidays(sheet, date_col_map, code_row_map):
    """For each public holiday date, fill 'Public Holiday' in in/out/shortage for all employees."""
    print("\nReading Public Holiday Report...")
    try:
        holidays_df = read_public_holidays()
    except FileNotFoundError:
        print("  Public Holiday report not found; skipping (optional).")
        return
    print(f"  {len(holidays_df)} public holiday(s) loaded.")

    filled = 0
    skipped_date = 0

    for _, row in holidays_df.iterrows():
        hol_date_val = row["Date of Public Holiday"]
        if isinstance(hol_date_val, pd.Timestamp):
            hol_date = hol_date_val.date()
        else:
            hol_date = datetime.strptime(str(hol_date_val).strip(), "%d/%m/%Y").date()

        if hol_date not in date_col_map:
            skipped_date += 1
            print(f"  Warning: Holiday date {hol_date} not found in Nagwa workday columns.")
            continue

        in_col = date_col_map[hol_date]
        out_col = in_col + 1
        shortage_col = in_col + 3

        for emp_code, nagwa_row in code_row_map.items():
            sheet.cell(nagwa_row, in_col).value = "Public Holiday"
            sheet.cell(nagwa_row, out_col).value = "Public Holiday"
            sheet.cell(nagwa_row, shortage_col).value = "Public Holiday"
            filled += 1

    print(f"\nPublic Holidays done! {filled} employee-day-cells filled.")
    if skipped_date:
        print(f"  {skipped_date} holiday date(s) not found in Nagwa sheet.")


def read_permissions():
    """Read the Permission Request report and return a cleaned DataFrame."""
    df = pd.read_excel(resolve_report_path(PERMISSIONS_BASENAME), engine="calamine")
    df.columns = df.columns.str.strip()
    if "Transaction Type" in df.columns:
        before_cancel = len(df)
        df = df[
            ~df["Transaction Type"].astype(str).str.contains(
                "cancel", case=False, na=False
            )
        ]
        excluded = before_cancel - len(df)
        if excluded:
            print(f"  {excluded} cancelled permission row(s) ignored.")
    if "Status" in df.columns:
        before_status = len(df)
        df = df[df["Status"].astype(str).str.strip() == "Approved"]
        excluded = before_status - len(df)
        if excluded:
            print(f"  {excluded} non-approved permission row(s) ignored.")
    df = df.dropna(subset=["Employee Code"])
    df["Employee Code"] = df["Employee Code"].astype(int)
    return df


LEAVE_ENTRY_SEPARATOR = " | "


def fill_permissions(sheet, date_col_map, code_row_map):
    """Fill the Leave column with 'Transaction Sub Type, Start Time, End Time' for each permission.

    When the source report contains more than one permission for the same
    employee/day, each entry is appended to the existing cell value using
    ``LEAVE_ENTRY_SEPARATOR`` so that downstream calculations can recover
    every individual entry.
    """
    print("\nReading Permissions Report...")
    permissions_df = read_permissions()
    print(f"  {len(permissions_df)} permission records loaded.")

    filled = 0
    appended = 0
    skipped_code = 0
    skipped_date = 0

    for _, row in permissions_df.iterrows():
        emp_code = int(row["Employee Code"])
        if emp_code not in code_row_map:
            skipped_code += 1
            continue

        date_val = row["Effective Date"]
        if isinstance(date_val, pd.Timestamp):
            eff_date = date_val.date()
        else:
            try:
                eff_date = datetime.strptime(str(date_val).strip(), "%d/%m/%Y").date()
            except ValueError:
                skipped_date += 1
                continue

        if eff_date not in date_col_map:
            skipped_date += 1
            continue

        sub_type = str(row["Transaction Sub Type"]).strip() if pd.notna(row["Transaction Sub Type"]) else ""
        start_time = str(row["Start Time"]).strip() if pd.notna(row["Start Time"]) else ""
        end_time = str(row["End Time"]).strip() if pd.notna(row["End Time"]) else ""
        leave_value = f"{sub_type}, {start_time}, {end_time}"

        nagwa_row = code_row_map[emp_code]
        leave_col = date_col_map[eff_date] + 2
        existing = sheet.cell(nagwa_row, leave_col).value
        existing_str = "" if existing is None else str(existing).strip()
        if existing_str in ("", "nan", "None"):
            sheet.cell(nagwa_row, leave_col).value = leave_value
            filled += 1
        else:
            sheet.cell(nagwa_row, leave_col).value = (
                f"{existing_str}{LEAVE_ENTRY_SEPARATOR}{leave_value}"
            )
            appended += 1

    print(f"\nPermissions done! {filled} cells filled, {appended} additional entries appended.")
    if skipped_code:
        print(f"  {skipped_code} rows skipped (employee code not found in Nagwa sheet).")
    if skipped_date:
        print(f"  {skipped_date} rows skipped (date not found in Nagwa sheet).")


def split_leave_entries(leave_value):
    """Split a Leave cell value into individual raw entry strings.

    Multiple permissions on the same day are joined with
    ``LEAVE_ENTRY_SEPARATOR``; this helper returns the list of trimmed,
    non-empty entries. Empty / blank cells yield an empty list.
    """
    if leave_value is None:
        return []
    raw = str(leave_value).strip()
    if raw in ("", "nan", "None"):
        return []
    return [p.strip() for p in raw.split(LEAVE_ENTRY_SEPARATOR) if p.strip()]


def parse_time_value(time_str):
    """Parse a time string flexibly into datetime.time.

    Handles 'h:mm AM/PM', 'hh:mm AM/PM', 'h:mmAM/PM', and 24-hour 'H:MM' / 'HH:MM'.
    Returns None when the value is empty, non-string, or unparseable.
    """
    if time_str is None:
        return None
    s = str(time_str).strip()
    if s in ("", "nan", "None"):
        return None
    for fmt in ("%I:%M %p", "%I:%M%p", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


def parse_leave_cell(leave_value):
    """Parse a Leave cell ('Type, Start Time, End Time') into a 3-tuple.

    Returns (leave_type, start_time, end_time) where times are datetime.time.
    Returns (None, None, None) for empty / blank cells.
    Raises ValueError when the format is malformed or times are unparseable.
    """
    if leave_value is None:
        return None, None, None
    raw = str(leave_value).strip()
    if raw in ("", "nan", "None"):
        return None, None, None
    parts = [p.strip() for p in raw.split(",")]
    if len(parts) < 3:
        raise ValueError(
            f"Expected at least 3 comma-separated parts, got {len(parts)}: '{raw}'"
        )
    end_str = parts[-1]
    start_str = parts[-2]
    leave_type = ", ".join(parts[:-2])
    start_time = parse_time_value(start_str)
    end_time = parse_time_value(end_str)
    if start_time is None:
        raise ValueError(f"Cannot parse leave start time: '{start_str}'")
    if end_time is None:
        raise ValueError(f"Cannot parse leave end time: '{end_str}'")
    return leave_type, start_time, end_time


def _infer_halfday_position(leave_start, leave_end):
    """Return 'first' or 'second' for a generic half-day leave.

    A leave whose off-block starts in the morning (before noon) means the
    first half of the day is off; otherwise the second half is off. Returns
    None when the times are unavailable, so the caller can fall back to no
    override rather than guessing.
    """
    if leave_start is None or leave_end is None:
        return None
    midday = _time_to_min(time(12, 0))
    return "first" if _time_to_min(leave_start) < midday else "second"


def classify_leave_type(leave_type, leave_start=None, leave_end=None):
    """Map a leave-type string to a rule category.

    Returns one of '1st_half_annual', '2nd_half_annual',
    '1st_half_ramadan', '2nd_half_ramadan', 'permission', 'work_mission',
    or None if the type does not trigger a shortage override.

    Generic "Half Day Annual"/"Half Day Ramadan" entries (e.g. the
    "Half Day Annual (Daylight Saving)" label) do not state which half is
    off, so the 1st/2nd half is inferred from ``leave_start``/``leave_end``.
    """
    normalized = re.sub(r"\s+", " ", leave_type.strip().lower())
    if "1st half day" in normalized and "annual" in normalized:
        return "1st_half_annual"
    if "2nd half day" in normalized and "annual" in normalized:
        return "2nd_half_annual"
    if "1st half day" in normalized and "ramadan" in normalized:
        return "1st_half_ramadan"
    if "2nd half day" in normalized and "ramadan" in normalized:
        return "2nd_half_ramadan"
    if "half day" in normalized and "annual" in normalized:
        position = _infer_halfday_position(leave_start, leave_end)
        if position == "first":
            return "1st_half_annual"
        if position == "second":
            return "2nd_half_annual"
        return None
    if "half day" in normalized and "ramadan" in normalized:
        position = _infer_halfday_position(leave_start, leave_end)
        if position == "first":
            return "1st_half_ramadan"
        if position == "second":
            return "2nd_half_ramadan"
        return None
    if normalized == "work mission":
        return "work_mission"
    if normalized == "permission":
        return "permission"
    if "educational leave" in normalized and "core of business" in normalized:
        return "permission"
    return None


def is_permitted_delay_type(leave_type):
    return "permitted delays" in str(leave_type or "").strip().lower()


def _time_to_min(t):
    """Convert datetime.time to minutes since midnight."""
    return t.hour * 60 + t.minute


def _overlap_minutes(pres_start, pres_end, win_start, win_end):
    """Minutes of overlap between an employee's presence and a required window."""
    lo = max(_time_to_min(pres_start), _time_to_min(win_start))
    hi = min(_time_to_min(pres_end), _time_to_min(win_end))
    return max(0, hi - lo)


def _merged_covered_minutes(intervals, win_start, win_end):
    """Return unique covered minutes inside [win_start, win_end]."""
    win_lo = _time_to_min(win_start)
    win_hi = _time_to_min(win_end)
    clipped = []
    for start, end in intervals:
        if start is None or end is None:
            continue
        lo = max(_time_to_min(start), win_lo)
        hi = min(_time_to_min(end), win_hi)
        if hi > lo:
            clipped.append((lo, hi))
    clipped.sort()

    covered = 0
    cur_lo = cur_hi = None
    for lo, hi in clipped:
        if cur_hi is None or lo > cur_hi:
            if cur_hi is not None:
                covered += cur_hi - cur_lo
            cur_lo, cur_hi = lo, hi
        else:
            cur_hi = max(cur_hi, hi)
    if cur_hi is not None:
        covered += cur_hi - cur_lo
    return covered


def _apply_halfday_rule(
    sheet, nagwa_row, in_col, shortage_col, category, leave_start, leave_end,
    in_time, out_time, emp_code, d, warnings, other_leaves=None,
):
    """Override the shortage cell for a single half-day-style leave entry.

    Returns True when the shortage cell was changed (or marked as Missing
    Punch), False otherwise. ``in_time`` / ``out_time`` are pre-parsed.

    ``other_leaves`` is an iterable of ``(start_time, end_time)`` tuples for
    every other (non-half-day) leave entry on the same day. Their overlap
    with the required half-day window is added to the employee's punch
    coverage, so e.g. a Work Mission spanning the rest of the morning fully
    covers the 2nd-half-day-annual window.
    """
    if in_time is None or out_time is None:
        if in_time is None:
            sheet.cell(nagwa_row, in_col).value = "Missing Punch"
        if out_time is None:
            sheet.cell(nagwa_row, in_col + 1).value = "Missing Punch"
        sheet.cell(nagwa_row, shortage_col).value = "Missing Punch"
        return True

    if category == "1st_half_annual":
        win_start, win_end = leave_end, WORKDAY_END_NORMAL
    elif category == "2nd_half_annual":
        win_start, win_end = WORKDAY_START, leave_start
    elif category == "1st_half_ramadan":
        win_start, win_end = leave_end, WORKDAY_END_RAMADAN
    else:
        win_start, win_end = WORKDAY_START, leave_start

    window_min = _time_to_min(win_end) - _time_to_min(win_start)
    if window_min <= 0:
        warnings.append(
            f"Row {nagwa_row} (Employee {emp_code}), Date {d}: "
            f"Required-presence window is zero or negative "
            f"({win_start} -> {win_end})"
        )
        return False

    intervals = [(_time_to_min(in_time), _time_to_min(out_time))]
    for o_start, o_end in (other_leaves or ()):
        if o_start is None or o_end is None:
            continue
        intervals.append((_time_to_min(o_start), _time_to_min(o_end)))

    win_lo = _time_to_min(win_start)
    win_hi = _time_to_min(win_end)
    clipped = []
    for lo, hi in intervals:
        lo_c = max(lo, win_lo)
        hi_c = min(hi, win_hi)
        if hi_c > lo_c:
            clipped.append((lo_c, hi_c))
    clipped.sort()
    covered = 0
    cur_lo = cur_hi = None
    for lo, hi in clipped:
        if cur_hi is None or lo > cur_hi:
            if cur_hi is not None:
                covered += cur_hi - cur_lo
            cur_lo, cur_hi = lo, hi
        else:
            cur_hi = max(cur_hi, hi)
    if cur_hi is not None:
        covered += cur_hi - cur_lo

    shortage = max(0, window_min - covered)
    sheet.cell(nagwa_row, shortage_col).value = minutes_to_hhmm(shortage)
    return True


def _apply_permission_rule(
    sheet, nagwa_row, in_col, shortage_col, leave_start, leave_end,
    in_time, out_time, d, window_end,
):
    """Override the shortage cell for a single permission-style entry.

    Mirrors the original single-entry permission logic. Returns True when
    the shortage cell was changed (or marked Missing Punch).

    Presence is clamped to the permitted attendance window
    [WORKDAY_START, window_end] to match ``calculate_shortage``: minutes
    before 8:00 AM and after ``window_end`` do not count as actual work.
    """
    full_day = FULL_DAY_RAMADAN if RAMADAN_START <= d <= RAMADAN_END else FULL_DAY_NORMAL
    window_start_min = _time_to_min(WORKDAY_START)
    window_end_min = _time_to_min(window_end)

    if in_time is None:
        sheet.cell(nagwa_row, in_col).value = "Missing Punch"
        if out_time is None:
            sheet.cell(nagwa_row, in_col + 1).value = "Missing Punch"
        sheet.cell(nagwa_row, shortage_col).value = "Missing Punch"
        return True

    leave_start_min = _time_to_min(leave_start)
    leave_end_min = _time_to_min(leave_end)

    if leave_start_min > _time_to_min(in_time):
        if out_time is None:
            sheet.cell(nagwa_row, in_col + 1).value = "Missing Punch"
            sheet.cell(nagwa_row, shortage_col).value = "Missing Punch"
            return True
        if leave_end_min <= _time_to_min(out_time):
            return False
        leave_dur = leave_end_min - leave_start_min
        effective_in = max(_time_to_min(in_time), window_start_min)
        effective_out = min(_time_to_min(out_time), window_end_min)
        actual_work = max(0, effective_out - effective_in)
        # Exclude any presence that overlaps the permission window so the
        # same minutes are not credited twice (once via the leave duration
        # and once via the punch interval).
        overlap = max(0, min(effective_out, leave_end_min) - max(effective_in, leave_start_min))
        actual_work -= overlap
        hours_shortage = max(0, (full_day - leave_dur) - actual_work)
        # Late sign-in after 10:00 AM is not covered by a leave that starts
        # later in the day, so any minutes past 10:00 still count as shortage.
        ten_am = _time_to_min(time(10, 0))
        late_penalty = max(0, _time_to_min(in_time) - ten_am)
        shortage = max(late_penalty, hours_shortage)
        sheet.cell(nagwa_row, shortage_col).value = minutes_to_hhmm(shortage)
        return True

    if out_time is None:
        sheet.cell(nagwa_row, in_col + 1).value = "Missing Punch"
        sheet.cell(nagwa_row, shortage_col).value = "Missing Punch"
        return True

    ten_am = _time_to_min(time(10, 0))
    cutoff = max(ten_am, leave_end_min)
    late_penalty = max(0, _time_to_min(in_time) - cutoff)

    leave_dur = leave_end_min - leave_start_min
    required_presence = full_day - leave_dur
    effective_in = max(_time_to_min(in_time), window_start_min)
    effective_out = min(_time_to_min(out_time), window_end_min)
    actual_presence = max(0, effective_out - effective_in)
    # Exclude any presence that overlaps the permission window so the same
    # minutes are not credited twice (once via the leave duration and once
    # via the punch interval).
    overlap = max(0, min(effective_out, leave_end_min) - max(effective_in, leave_start_min))
    actual_presence -= overlap
    hours_shortage = max(0, required_presence - actual_presence)

    total_shortage = max(late_penalty, hours_shortage)
    sheet.cell(nagwa_row, shortage_col).value = minutes_to_hhmm(total_shortage)
    return True


def _apply_work_mission_rule(
    sheet, nagwa_row, shortage_col, mission_intervals, in_time, out_time, d, window_end,
):
    """Treat Work Mission intervals as worked/covered time on their effective day."""
    if d and RAMADAN_START <= d <= RAMADAN_END:
        full_day = FULL_DAY_RAMADAN
    else:
        full_day = FULL_DAY_NORMAL

    intervals = list(mission_intervals)
    if in_time is not None and out_time is not None:
        intervals.append((in_time, out_time))

    covered = _merged_covered_minutes(intervals, WORKDAY_START, window_end)
    shortage = max(0, full_day - covered)
    sheet.cell(nagwa_row, shortage_col).value = minutes_to_hhmm(shortage)
    return True


def _subtract_leave_from_shortage(sheet, nagwa_row, shortage_col, leave_start, leave_end):
    """Subtract a leave entry's duration from the current shortage value.

    Used for permission-style entries that come AFTER another rule-bearing
    entry has already redefined the shortage on the same day. Cells holding
    non-numeric statuses (e.g. 'Missing Punch') are left untouched.
    Returns True when the shortage cell was changed.
    """
    leave_dur = _time_to_min(leave_end) - _time_to_min(leave_start)
    if leave_dur <= 0:
        return False
    current = sheet.cell(nagwa_row, shortage_col).value
    if current is None:
        return False
    current_str = str(current).strip()
    if not re.match(r"^\d+:\d{2}$", current_str):
        return False
    new_minutes = max(0, parse_duration(current_str) - leave_dur)
    sheet.cell(nagwa_row, shortage_col).value = minutes_to_hhmm(new_minutes)
    return True


_HALFDAY_CATEGORIES = (
    "1st_half_annual",
    "2nd_half_annual",
    "1st_half_ramadan",
    "2nd_half_ramadan",
)


def recalculate_shortage_from_leave(sheet, date_col_map, code_row_map, code_schedule_map=None):
    """Analyse the Leave column and conditionally override shortage values.

    Runs AFTER all existing processing (attendance, vacations, resignations,
    public holidays, permissions).  Only overrides shortage when a recognised
    leave-type rule explicitly applies; all other rows are left untouched.

    Multiple leave entries on the same day (separated by
    ``LEAVE_ENTRY_SEPARATOR``) are processed sequentially: the first
    rule-bearing entry redefines the shortage using its own rule, and each
    subsequent permission-style entry then subtracts its duration from the
    running shortage value. A later half-day-style entry re-redefines the
    shortage using its own window rule.
    """
    print("\nRecalculating shortage based on leave rules...")
    warnings = []
    overridden = 0
    sorted_dates = sorted(date_col_map.keys())

    for emp_code, nagwa_row in code_row_map.items():
        schedule = (code_schedule_map or {}).get(emp_code)
        for d in sorted_dates:
            in_col = date_col_map[d]
            leave_col = in_col + 2
            shortage_col = in_col + 3

            raw_entries = split_leave_entries(sheet.cell(nagwa_row, leave_col).value)
            if not raw_entries:
                continue

            in_time = parse_time_value(sheet.cell(nagwa_row, in_col).value)
            out_time = parse_time_value(sheet.cell(nagwa_row, in_col + 1).value)
            window_end = _permitted_window_end(d, emp_code, schedule)

            parsed_entries = []
            for raw_entry in raw_entries:
                try:
                    leave_type, leave_start, leave_end = parse_leave_cell(raw_entry)
                except ValueError as exc:
                    warnings.append(
                        f"Row {nagwa_row} (Employee {emp_code}), Date {d}: {exc}"
                    )
                    parsed_entries.append(None)
                    continue
                if leave_type is None:
                    parsed_entries.append(None)
                    continue
                category = classify_leave_type(leave_type, leave_start, leave_end)
                parsed_entries.append((leave_type, leave_start, leave_end, category))

            mission_intervals = [
                (p[1], p[2])
                for p in parsed_entries
                if p is not None and p[3] == "work_mission"
            ]
            has_halfday = any(
                p is not None and p[3] in _HALFDAY_CATEGORIES
                for p in parsed_entries
            )

            rule_applied = False
            any_change = False
            if mission_intervals and not has_halfday:
                if _apply_work_mission_rule(
                    sheet, nagwa_row, shortage_col, mission_intervals,
                    in_time, out_time, d, window_end,
                ):
                    rule_applied = True
                    any_change = True

            for idx, parsed in enumerate(parsed_entries):
                if parsed is None:
                    continue
                leave_type, leave_start, leave_end, category = parsed
                if category is None:
                    continue

                if category in _HALFDAY_CATEGORIES:
                    other_leaves = [
                        (p[1], p[2])
                        for j, p in enumerate(parsed_entries)
                        if (
                            j != idx
                            and p is not None
                            and p[3] not in _HALFDAY_CATEGORIES
                            and not is_permitted_delay_type(p[0])
                        )
                    ]
                    changed = _apply_halfday_rule(
                        sheet, nagwa_row, in_col, shortage_col, category,
                        leave_start, leave_end, in_time, out_time,
                        emp_code, d, warnings, other_leaves=other_leaves,
                    )
                    if changed:
                        rule_applied = True
                        any_change = True
                        in_time = parse_time_value(sheet.cell(nagwa_row, in_col).value)
                        out_time = parse_time_value(sheet.cell(nagwa_row, in_col + 1).value)
                elif category == "permission":
                    if not rule_applied:
                        changed = _apply_permission_rule(
                            sheet, nagwa_row, in_col, shortage_col,
                            leave_start, leave_end, in_time, out_time, d,
                            window_end,
                        )
                        if changed:
                            rule_applied = True
                            any_change = True
                            in_time = parse_time_value(sheet.cell(nagwa_row, in_col).value)
                            out_time = parse_time_value(sheet.cell(nagwa_row, in_col + 1).value)
                    else:
                        if has_halfday:
                            # The half-day rule already counted normal
                            # permissions as covered time inside its required
                            # work window. Do not subtract the same interval a
                            # second time; permitted delays are handled later.
                            continue
                        if _subtract_leave_from_shortage(
                            sheet, nagwa_row, shortage_col, leave_start, leave_end,
                        ):
                            any_change = True
                elif category == "work_mission":
                    continue

            if any_change:
                overridden += 1

    print(f"Shortage recalculation done! {overridden} shortage value(s) overridden.")
    if warnings:
        print(f"\n--- WARNING REPORT: {len(warnings)} issue(s) found ---")
        for w in warnings:
            print(f"  * {w}")
    else:
        print("No data warnings.")


def apply_permitted_delays(sheet, date_col_map, code_row_map):
    """Subtract permitted-delay duration from shortage as the final adjustment.

    For leave entries whose type contains 'permitted delays', the duration
    (end_time - start_time) is subtracted from the current shortage value.
    Negative results are clamped to zero.
    """
    print("\nApplying permitted-delay deductions...")
    adjusted = 0
    warnings = []
    sorted_dates = sorted(date_col_map.keys())

    for emp_code, nagwa_row in code_row_map.items():
        for d in sorted_dates:
            in_col = date_col_map[d]
            leave_col = in_col + 2
            shortage_col = in_col + 3

            raw_entries = split_leave_entries(sheet.cell(nagwa_row, leave_col).value)
            if not raw_entries:
                continue

            for raw_entry in raw_entries:
                try:
                    leave_type, leave_start, leave_end = parse_leave_cell(raw_entry)
                except ValueError:
                    continue

                if leave_type is None:
                    continue

                if "permitted delays" not in leave_type.lower():
                    continue

                delay_minutes = _time_to_min(leave_end) - _time_to_min(leave_start)
                if delay_minutes <= 0:
                    warnings.append(
                        f"Row {nagwa_row} (Employee {emp_code}), Date {d}: "
                        f"Permitted-delay duration is zero or negative "
                        f"({leave_start} -> {leave_end})"
                    )
                    continue

                current_shortage = sheet.cell(nagwa_row, shortage_col).value
                current_str = "" if current_shortage is None else str(current_shortage).strip()
                if not re.match(r"^\d+:\d{2}$", current_str):
                    continue
                current_minutes = parse_duration(current_str)

                new_shortage = max(0, current_minutes - delay_minutes)
                sheet.cell(nagwa_row, shortage_col).value = minutes_to_hhmm(new_shortage)
                adjusted += 1

    print(f"Permitted-delay deductions done! {adjusted} shortage value(s) adjusted.")
    if warnings:
        print(f"\n--- WARNING REPORT: {len(warnings)} issue(s) found ---")
        for w in warnings:
            print(f"  * {w}")


def apply_work_mission_lunch_exemption(sheet, date_col_map, code_row_map):
    """Exempt the lunch hour from shortage on Work-Mission days.

    For any day where the Leave column contains at least one Work Mission
    entry, the relevant lunch window (12:00 PM - 1:00 PM before 26 Apr 2026,
    12:30 PM - 1:30 PM from that date on) is treated as covered: minutes
    inside the window that are NOT already covered by the employee's punch
    interval or by any leave entry are subtracted from the current shortage.
    Non-numeric shortage cells ('Missing Punch', 'Public Holiday', vacation
    types, etc.) are left untouched.
    """
    print("\nApplying Work-Mission lunch-gap exemption...")
    adjusted = 0

    for nagwa_row in code_row_map.values():
        for d, in_col in date_col_map.items():
            leave_col = in_col + 2
            shortage_col = in_col + 3

            raw_entries = split_leave_entries(sheet.cell(nagwa_row, leave_col).value)
            if not raw_entries:
                continue

            has_work_mission = False
            covered_intervals = []
            for raw_entry in raw_entries:
                try:
                    leave_type, leave_start, leave_end = parse_leave_cell(raw_entry)
                except ValueError:
                    continue
                if leave_type is None or leave_start is None or leave_end is None:
                    continue
                if "work mission" in leave_type.lower():
                    has_work_mission = True
                covered_intervals.append(
                    (_time_to_min(leave_start), _time_to_min(leave_end))
                )

            if not has_work_mission:
                continue

            current_shortage = sheet.cell(nagwa_row, shortage_col).value
            if current_shortage is None:
                continue
            shortage_str = str(current_shortage).strip()
            if not re.match(r"^\d+:\d{2}$", shortage_str):
                continue

            in_time = parse_time_value(sheet.cell(nagwa_row, in_col).value)
            out_time = parse_time_value(sheet.cell(nagwa_row, in_col + 1).value)
            if in_time is not None and out_time is not None:
                covered_intervals.append(
                    (_time_to_min(in_time), _time_to_min(out_time))
                )

            lunch_start, lunch_end = _lunch_window_for(d)
            win_lo = _time_to_min(lunch_start)
            win_hi = _time_to_min(lunch_end)
            window_minutes = win_hi - win_lo
            if window_minutes <= 0:
                continue

            clipped = []
            for lo, hi in covered_intervals:
                lo_c = max(lo, win_lo)
                hi_c = min(hi, win_hi)
                if hi_c > lo_c:
                    clipped.append((lo_c, hi_c))
            clipped.sort()
            covered = 0
            cur_lo = cur_hi = None
            for lo, hi in clipped:
                if cur_hi is None or lo > cur_hi:
                    if cur_hi is not None:
                        covered += cur_hi - cur_lo
                    cur_lo, cur_hi = lo, hi
                else:
                    cur_hi = max(cur_hi, hi)
            if cur_hi is not None:
                covered += cur_hi - cur_lo

            uncovered = window_minutes - covered
            if uncovered <= 0:
                continue

            current_minutes = parse_duration(shortage_str)
            new_minutes = max(0, current_minutes - uncovered)
            if new_minutes == current_minutes:
                continue
            sheet.cell(nagwa_row, shortage_col).value = minutes_to_hhmm(new_minutes)
            adjusted += 1

    print(f"Work-Mission lunch exemption done! {adjusted} shortage value(s) adjusted.")


def apply_hour_reduction(sheet, date_col_map, code_row_map):
    """Subtract 1 hour from shortage for specific employees within their windows.

    For each employee listed in ``HOUR_REDUCTION_WINDOWS``, every numeric
    shortage value on a date inside the configured (inclusive) range is
    reduced by 60 minutes; values below an hour become 0:00. Non-numeric
    shortages (e.g. 'Missing Punch', 'Public Holiday', vacation types) are
    left untouched.
    """
    print("\nApplying 1-hour shortage reductions for specific employees...")
    adjusted = 0

    for emp_code, (start_date, end_date) in HOUR_REDUCTION_WINDOWS.items():
        if emp_code not in code_row_map:
            continue
        nagwa_row = code_row_map[emp_code]

        for d, in_col in date_col_map.items():
            if start_date is not None and d < start_date:
                continue
            if end_date is not None and d > end_date:
                continue

            shortage_col = in_col + 3
            current_shortage = sheet.cell(nagwa_row, shortage_col).value
            if current_shortage is None:
                continue
            shortage_str = str(current_shortage).strip()
            if shortage_str in ("", "nan", "None"):
                continue
            if not re.match(r"^\d+:\d{2}$", shortage_str):
                continue

            current_minutes = parse_duration(shortage_str)
            new_minutes = max(0, current_minutes - 60)
            sheet.cell(nagwa_row, shortage_col).value = minutes_to_hhmm(new_minutes)
            adjusted += 1

    print(f"1-hour shortage reductions done! {adjusted} shortage value(s) adjusted.")


def fill_missing_punches(sheet, date_col_map, code_row_map, schedule_map=None, status_map=None):
    """Final sweep: where exactly one of in/out is empty but the other holds a
    real punch time, treat the day as absent.

    Earlier leave-rule steps may also mark a row as 'Missing Punch'. Those
    cases are normalized to 'absent' here so the final report receives the
    standard A flag.

    Employees with Schedule = Undefined or Employment Status = Challenged 5%
    are skipped (not eligible for absence calculations).
    """
    print("\nScanning for missing punches to mark as absences...")
    filled = 0
    schedule_map = schedule_map or {}
    status_map = status_map or {}

    for emp_code, nagwa_row in code_row_map.items():
        if is_absence_exempt(emp_code, schedule_map, status_map):
            continue
        for d, in_col in date_col_map.items():
            out_col = in_col + 1
            leave_col = in_col + 2
            shortage_col = in_col + 3

            in_val = sheet.cell(nagwa_row, in_col).value
            out_val = sheet.cell(nagwa_row, out_col).value
            leave_val = sheet.cell(nagwa_row, leave_col).value

            in_str = "" if in_val is None else str(in_val).strip()
            out_str = "" if out_val is None else str(out_val).strip()
            shortage_str = "" if sheet.cell(nagwa_row, shortage_col).value is None else str(sheet.cell(nagwa_row, shortage_col).value).strip()
            leave_str = "" if leave_val is None else str(leave_val).strip()

            if "work mission" in leave_str.lower() and re.match(r"^\d+:\d{2}$", shortage_str):
                continue

            if "Missing Punch" in (in_str, out_str, shortage_str):
                sheet.cell(nagwa_row, in_col).value = "absent"
                sheet.cell(nagwa_row, out_col).value = "absent"
                sheet.cell(nagwa_row, shortage_col).value = "absent"
                filled += 1
                continue

            in_blank = in_str in ("", "nan", "None")
            out_blank = out_str in ("", "nan", "None")

            if in_blank == out_blank:
                continue

            if in_blank:
                if parse_time_value(out_str) is None:
                    continue
            else:
                if parse_time_value(in_str) is None:
                    continue

            sheet.cell(nagwa_row, in_col).value = "absent"
            sheet.cell(nagwa_row, out_col).value = "absent"
            sheet.cell(nagwa_row, shortage_col).value = "absent"
            filled += 1

    print(f"Missing punch absence sweep done! {filled} day-cell(s) marked absent.")


def _cell_is_blank(value):
    """True if a sheet cell is empty (None or a blank/placeholder string)."""
    if value is None:
        return True
    return str(value).strip() in ("", "nan", "None")


def _is_employed_on(employment_map, emp_code, att_date):
    """True if ``emp_code`` is within its [hire, termination] window on a date.

    A missing hire date is treated as "employed from the beginning" and a
    missing termination date as "still employed", so only positively-known
    pre-hire / post-termination days are excluded.
    """
    hire, term = employment_map.get(emp_code, (None, None))
    if hire is not None and att_date < hire:
        return False
    if term is not None and att_date > term:
        return False
    return True


def fill_full_day_absences(
    sheet, date_col_map, code_row_map, employment_map,
    coverage_start=None, coverage_end=None,
    schedule_map=None, status_map=None,
):
    """Closed-world safety net: mark scheduled working days that finished the
    pipeline completely empty as 'absent'.

    This backs up the Absence Report so a full no-show is still caught even if
    the report omits it. A cell is only marked when ALL of these hold:

      * the date is a scheduled working-day column (weekends are single-cell
        columns and never appear in ``date_col_map``);
      * the date is inside the reports' coverage window
        ``[coverage_start, coverage_end]``, so days the period covers but the
        source data does not report yet are never treated as absences;
      * the in / out / Leave / Shortage cells are all blank, so anything
        already resolved - punches, leave, holiday, vacation, resignation,
        WFH/Workday, permission shortages - is left untouched;
      * the employee was actively employed on that date (hire/termination
        window from the Nagwa sheet), so a new hire's pre-hire days and a
        leaver's post-termination days are not misreported as absences;
      * the employee is not absence-exempt (Schedule = Undefined or
        Employment Status = Challenged 5%).

    Every cell marked here is, by definition, a working-day absence that was
    NOT captured by the Absence Report, so each one is logged for review.
    """
    print("\nScanning for uncaptured full-day absences...")
    filled = 0
    flagged = []
    schedule_map = schedule_map or {}
    status_map = status_map or {}

    for emp_code, nagwa_row in code_row_map.items():
        if is_absence_exempt(emp_code, schedule_map, status_map):
            continue
        for att_date, in_col in date_col_map.items():
            if coverage_start is not None and att_date < coverage_start:
                continue
            if coverage_end is not None and att_date > coverage_end:
                continue
            all_blank = all(
                _cell_is_blank(sheet.cell(nagwa_row, in_col + k).value)
                for k in range(4)
            )
            if not all_blank:
                continue
            if not _is_employed_on(employment_map, emp_code, att_date):
                continue

            sheet.cell(nagwa_row, in_col).value = "absent"
            sheet.cell(nagwa_row, in_col + 1).value = "absent"
            sheet.cell(nagwa_row, in_col + 3).value = "absent"
            filled += 1
            flagged.append((emp_code, att_date))

    print(f"Full-day absence sweep done! {filled} day-cell(s) marked absent.")
    if flagged:
        print(
            f"\n--- WARNING REPORT: {len(flagged)} full-day absence(s) not "
            f"present in the Absence Report ---"
        )
        for emp_code, att_date in sorted(flagged, key=lambda pair: (pair[0], pair[1])):
            print(
                f"  * Employee {emp_code}, Date {att_date}: no attendance/leave "
                f"recorded on a scheduled working day; marked absent."
            )


def apply_wfh_and_workday_overrides(sheet, date_col_map, code_row_map):
    """Final override: if the Leave cell mentions 'Work from Home' or 'Workday',
    fill the in/out/shortage cells with 'WFH' or 'Workday' respectively.

    'Workday' takes precedence when both keywords appear in the same cell.
    """
    print("\nApplying WFH / Workday leave overrides...")
    filled = 0

    for nagwa_row in code_row_map.values():
        for d, in_col in date_col_map.items():
            out_col = in_col + 1
            leave_col = in_col + 2
            shortage_col = in_col + 3

            leave_val = sheet.cell(nagwa_row, leave_col).value
            if leave_val is None:
                continue
            leave_str = str(leave_val)
            leave_lower = leave_str.lower()

            if "workday" in leave_lower:
                label = "Workday"
            elif "work from home" in leave_lower:
                label = "WFH"
            else:
                continue

            sheet.cell(nagwa_row, in_col).value = label
            sheet.cell(nagwa_row, out_col).value = label
            sheet.cell(nagwa_row, shortage_col).value = label
            filled += 1

    print(f"WFH / Workday overrides done! {filled} day-cell(s) filled.")


if __name__ == "__main__":
    main()
